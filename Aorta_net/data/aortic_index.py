import SimpleITK as sitk
import os
import numpy as np
from skimage import measure, morphology
from scipy.spatial.distance import cdist
import h5py
import matplotlib.pyplot as plt
from skimage.morphology import binary_dilation
import copy
from mayavi import mlab
import cv2
from skimage.morphology import binary_closing
from scipy.stats import pearsonr,binned_statistic
import pingouin as pg
from sklearn.metrics import mean_absolute_error,r2_score,mean_absolute_percentage_error
from natsort import natsorted
import openpyxl
from sklearn.metrics import confusion_matrix
import pandas as pd
from scipy.ndimage import label, generate_binary_structure
# 设置打印选项，使得所有数组都以小数形式输出，且设置小数点后保留的位数
np.set_printoptions(suppress=True, precision=8)  # suppress=True 禁用科学记数法，precision设置小数点后的位数

#remain the bigest Connected region
def remove_regions(mask):
    mask1=mask
    mask1=np.where(mask1>0,1,0)
    # mask2 = mask
    # mask2=np.where(mask2>1,1,0)
    segmentation_sitk = sitk.GetImageFromArray(mask1)
    # 计算标签图像的连通分量
    connected_components_filter = sitk.ConnectedComponentImageFilter()
    labeled_image = connected_components_filter.Execute(segmentation_sitk)

    # 获取每个连通分量的大小
    label_shape_filter = sitk.LabelShapeStatisticsImageFilter()
    label_shape_filter.Execute(labeled_image)

    # 找到最大的连通分量ID
    max_size = 0
    largest_label = 0
    for i in range(1, label_shape_filter.GetNumberOfLabels() + 1):  # Label index starts from 1
        if label_shape_filter.GetNumberOfPixels(i) > max_size:
            max_size = label_shape_filter.GetNumberOfPixels(i)
            largest_label = i

    # 仅保留最大连通分量
    binary_mask = sitk.Equal(labeled_image, largest_label)
    cleaned_segmentation = sitk.Cast(binary_mask, segmentation_sitk.GetPixelID())
    cleaned_segmentation = sitk.GetArrayFromImage(cleaned_segmentation)
    cleaned_segmentation=cleaned_segmentation*mask
    # print(cleaned_segmentation.max())
    return cleaned_segmentation.astype(np.int16)

##remain the  Connected region whichs more than 1000 voxel
def remove_small_volums(mask):
    mask1 = np.where(mask > 0, 1, 0)
    segmentation_sitk = sitk.GetImageFromArray(mask1)
    # 计算标签图像的连通分量
    connected_components_filter = sitk.ConnectedComponentImageFilter()
    labeled_image = connected_components_filter.Execute(segmentation_sitk)
    # 获取每个连通分量的大小
    label_shape_filter = sitk.LabelShapeStatisticsImageFilter()
    label_shape_filter.Execute(labeled_image)
    # 初始化一个空的数组来存储处理后的mask
    cleaned_segmentation = np.zeros_like(mask1)
    # 遍历每个连通分量，保留体积大于min_volume的连通区域
    for i in range(1, label_shape_filter.GetNumberOfLabels() + 1):
        if label_shape_filter.GetNumberOfPixels(i) >= 14000:
            binary_mask = sitk.Equal(labeled_image, i)
            binary_mask_array = sitk.GetArrayFromImage(binary_mask)
            cleaned_segmentation[binary_mask_array == 1] = 1
    # 返回处理后的mask
    cleaned_segmentation = cleaned_segmentation * mask
    return cleaned_segmentation.astype(np.int16)

def my_mape1(y_true, y_pred):
    non_zero_mask = np.nonzero(y_true)
    non_zero_maskk=np.nonzero(y_pred)## 找出y_pred中为非0的位置，避免直接除以0
    with np.errstate(invalid='ignore'):
        relative_errors_non_zero = np.abs((y_true[non_zero_mask] - y_pred[non_zero_mask]) / y_true[non_zero_mask]) * 100.0
    relative_errors = np.zeros_like(y_true)  # 初始化一个全nan组
    relative_errors[non_zero_maskk]=100#
    relative_errors[non_zero_mask] = relative_errors_non_zero  # 填充非零位置的计算结果
    errors=np.clip(relative_errors, 0, 100)#Winsorization（温索化处理）边界处理方法
    merged_indices = np.nonzero(relative_errors)
    return errors,merged_indices

def my_mape(y_true, y_pred):
    if y_true==0 and y_pred!=0:
        error=100
    elif y_pred==0 and y_true!=0:
        error=0
    elif y_pred==0 and y_true==0:
        error=0
    else:
        error = np.abs((y_true - y_pred) / y_true) * 100.0
        error = np.clip(error, 0, 100)
    return error

# def remove_ascend(image_stack):
#     # 定义结构元素用于确定连通性（这里假定是8邻域）
#     struct = generate_binary_structure(2, 2)
#
#     # 对每个切片进行连通组件分析
#     ascend_index=[]
#     i=0
#     labeled_slices = []
#     for slice_lume in image_stack:
#         img_label, num = measure.label(slice_lume, connectivity=2, return_num=True)
#         props = measure.regionprops(img_label)
#         props_sorted = sorted(props, key=lambda x: x.area, reverse=True)
#         if len(props_sorted) > 1:
#             if props_sorted[1].area>600:
#                 a=1
#
#         if len(props_sorted) > 1 and props_sorted[0].area > 600:
#             # max_label = props_sorted[1].label
#             if props_sorted[0].area < 3 * (props_sorted[1].area):
#                 max_label = props_sorted[1].label
#             else:
#                 max_label = props_sorted[0].label
#         else:
#             max_label = props_sorted[0].label
#         i=i+1
#     # 返回过滤后的3D mask
#     return np.stack(labeled_slices)

def diameter_area(mask):
    # lumen_mask = np.where(mask == 1, 1, 0)  # 管腔（正常血液流动区域）
    z_dims = mask.shape[0]
    area_per_lumen = []
    diameter_per_lumen = []
    flag_area=np.ones([512,512])
    for slice_idx in range(z_dims):  # 获取当前切片
        slice_lume = mask[slice_idx, :, :]
        if slice_lume.sum()==0:
            diameter_per_lumen.append(0)
            area_per_lumen.append(0)  # 将面积添加到列表中
        else:
            img_label, num = measure.label(slice_lume, connectivity=2, return_num=True)
            props = measure.regionprops(img_label)
            props_sorted = sorted(props, key=lambda x: x.area, reverse=True)
            if len(props_sorted) > 1 and props_sorted[0].area>300 and slice_idx>0:
                if props_sorted[0].area < 3 * (props_sorted[1].area):
                    max_label0 = props_sorted[0].label
                    slice_lume0 = (img_label == max_label0).astype(int)
                    overlap0 = np.logical_and(slice_lume0, flag_area)
                    overlap_area0 = np.count_nonzero(overlap0)

                    max_label = props_sorted[1].label
                    slice_lume = (img_label == max_label).astype(int)
                    overlap = np.logical_and(slice_lume, flag_area)
                    overlap_area = np.count_nonzero(overlap)

                    # if overlap_area < 0.2 * slice_lume.sum() or overlap_area < 0.33 * flag_area.sum():
                    if overlap_area < 0.33 * flag_area.sum() or overlap_area<0.66*overlap_area0:
                        max_label = props_sorted[0].label
                else:
                    max_label = props_sorted[0].label
                    slice_lume = (img_label == max_label).astype(int)
                    overlap = np.logical_and(slice_lume, flag_area)
                    overlap_area = np.count_nonzero(overlap)
                    if overlap_area<0.2*slice_lume.sum() or overlap_area<0.33*flag_area.sum():
                    # if overlap_area < 0.33 * flag_area.sum():
                        max_label = props_sorted[1].label
            else:
                max_label = props_sorted[0].label
                # slice_lume = (img_label == max_label).astype(int)
            # if slice_idx>310:
            #     a=1
            slice_lume = (img_label == max_label).astype(int)
            flag_area=slice_lume
            filled_slice_lume = binary_closing(slice_lume)
            gray_img = np.uint8(filled_slice_lume * 255)
            contours, _ = cv2.findContours(gray_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            dist_map = cv2.distanceTransform(gray_img, cv2.DIST_L2, cv2.DIST_MASK_PRECISE)
            _, radius, _, center = cv2.minMaxLoc(dist_map)

            # if slice_idx>265:
            #     result = cv2.cvtColor(gray_img, cv2.COLOR_GRAY2BGR)
            #     cv2.circle(result, tuple(center), int(radius), (0, 0, 255), 2, cv2.LINE_8, 0)
            #     cv2.imshow("result", result)
            #     print(radius*2)
            #     cv2.waitKey(0)
            #     a=1

            diameter=2*radius
            # threshold = np.percentile(distances, 2)
            area = np.sum(slice_lume > 0)#计算填充前的面积
            if area > np.pi * (diameter / 2) ** 2:
                area = np.pi * diameter ** 2
            diameter_per_lumen.append(diameter)
            area_per_lumen.append(area)
    return diameter_per_lumen, area_per_lumen

def aortic_index(path,path_save):
    # 一：读取保存有主动脉mask的NIfTI文件
    # mask_sitk = sitk.ReadImage(path, sitk.sitkInt16)
    # mask_image = sitk.GetArrayFromImage(mask_sitk)
    read = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
    img_array = sitk.GetArrayFromImage(read)
    mask_image=copy.deepcopy(img_array)
    img_array[img_array>2]=3
    img_array = remove_small_volums(img_array)#去掉冗余  remove_small_volums
    # img_array = remove_regions(img_array)  # 去掉冗余  remove_small_volums
    out = sitk.GetImageFromArray(img_array)
    sitk.WriteImage(out, path_save)

    # mask_image = remove_ascend(mask_image)
    lumen_calcium_mask=copy.deepcopy(mask_image)
    lumen_calcium_mask[lumen_calcium_mask > 2] = 0
    lumen_calcium_mask=remove_small_volums(lumen_calcium_mask) #去除管腔碎片，动脉瘤部分的钙化不应该纳入进来
    lumen_calcium_per_area=np.where(lumen_calcium_mask > 0, 1, 0).sum(axis=(1, 2))
    # plaque_mask = np.where(mask_image >= 2, 1, 0)  # 包括钙化板块和软斑块在内的所有异常区域

    lumen_mask=copy.deepcopy(lumen_calcium_mask)#管腔
    lumen_mask[lumen_mask > 1] = 0
    diameter_per_lumen,area_per_lumen=diameter_area(lumen_mask)
    lumen_diameter_index = np.percentile(diameter_per_lumen, [25, 50, 75])
    lumen_area_index = np.percentile(area_per_lumen, [25, 50, 75])

    total_mask = np.where(mask_image >0, 1, 0)
    total_mask=remove_small_volums(total_mask)
    diameter_per_total,area_per_total=diameter_area(total_mask)
    total_diameter_index = np.percentile(diameter_per_total, [25, 50, 75])
    total_area_index = np.percentile(area_per_total, [25, 50, 75])

    # print(lumen_calcium_mask.max())
    calcium_area = np.where(lumen_calcium_mask == 2, 1, 0).sum(axis=(1, 2))  #钙化板块
    # print(calcium_area.max())
    if np.sum(calcium_area)>0:#存在钙化
        calcium_per_index=calcium_area/lumen_calcium_per_area#钙化指数应该是钙化面积/(管腔+钙化)
        calcium_per_index[np.isnan(calcium_per_index)] = 0
        calcium_per_index[np.isinf(calcium_per_index)] = 100
        calciun_index = calcium_per_index.tolist()
        calciun_index = [num for num in calciun_index if num != 0]
        total_calcium_index = np.percentile(calciun_index, [25, 50, 75])
    else:
        calcium_per_index=np.ones(mask_image.shape[0])*0
        calcium_per_index=calcium_per_index.tolist()
        total_calcium_index = np.percentile(calcium_per_index, [25, 50, 75])
    # file=path.replace(".nii.gz",'.h5')
    h5_path_save = path_save.replace(".nii.gz", ".h5")
    if os.path.exists(h5_path_save):
        os.remove(h5_path_save)

    with h5py.File(h5_path_save, 'w') as f:# 创建一个dataset
        f.create_dataset('diameter_per_lumen', data=diameter_per_lumen)#管腔
        f.create_dataset('area_per_lumen', data=area_per_lumen)
        f.create_dataset('lumen_diameter_index', data=lumen_diameter_index)#index 为四分位数
        f.create_dataset('lumen_area_index', data=lumen_area_index)

        f.create_dataset('diameter_per_total', data=diameter_per_total)#整体
        f.create_dataset('area_per_total', data=area_per_total)
        f.create_dataset('total_diameter_index', data=total_diameter_index)
        f.create_dataset('total_area_index', data=total_area_index)

        f.create_dataset('calcium_per_index', data=calcium_per_index)#钙化
        f.create_dataset('total_calcium_index', data=total_calcium_index)

        # max_len=mask_image.shape[0]
        # fig, axs = plt.subplots(nrows=1, ncols=3, figsize=(6, 6), sharey=True, layout='constrained')
        # line_color = 'b'
        # line_width = 1
        # q1_color ="deepskyblue"
        # q2_color = "r"
        # for ax in axs:
        #     ax.invert_yaxis()
        #
        #     # ax.set_xlim(bottom=0)  # 设置y轴最小值为0
        #     # ax.set_ylim(bottom=0)  # 设置y轴最小值为0
        #     # ax.grid(True)  # 启用网格
        #
        # # axs[0].invert_yaxis()
        # # axs[0].grid(True)
        # axs[0].plot(diameter_per_lumen[::-1], range(max_len), color=line_color, linewidth=line_width)
        # axs[0].axvline(x=lumen_diameter_index[0], color=q1_color, linestyle='--')
        # axs[0].axvline(x=lumen_diameter_index[1], color=q2_color, linestyle='--')
        # axs[0].axvline(x=lumen_diameter_index[2], color=q1_color, linestyle='--')
        # axs[0].set_title('Lumen Diameter')
        #
        # # axs[1].invert_yaxis()
        # # axs[1].grid(True)
        # axs[1].plot(diameter_per_total[::-1], range(max_len), color=line_color, linewidth=line_width)
        # axs[1].axvline(x=total_diameter_index[0], color=q1_color, linestyle='--')
        # axs[1].axvline(x=total_diameter_index[1], color=q2_color, linestyle='--')
        # axs[1].axvline(x=total_diameter_index[2], color=q1_color, linestyle='--')
        # axs[1].set_title('Vessel Diameter')
        #
        # # axs[2].invert_yaxis()
        # # axs[2].grid(True)
        # axs[2].plot(calcium_per_index[::-1],range(max_len), color=line_color, linewidth=line_width)
        # axs[2].axvline(x=total_calcium_index[0], color=q1_color, linestyle='--')
        # axs[2].axvline(x=total_calcium_index[1], color=q2_color, linestyle='--')
        # axs[2].axvline(x=total_calcium_index[2], color=q1_color, linestyle='--')
        # axs[2].set_title('Calcium Index')
        # plt.show()

def Aortic_index_caculate():
    path = "/media/bit301/data/yml/data/p2_nii/external/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0
    for path in path_list:
        # if i<270:
        #     i=i+1
        #     continue
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dml/PA14/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA199/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/cq/nor/PA276/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/hnnk/dis/dmzyyh/DICOM0/PA14/SE0/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/hnnk/nor/PA3/SE0/2.nii.gz"
        # path="/media/bit301/data/yml/data/p2_nii/external/lz/dis/dml/PA10/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA117/2.nii.gz"
        # path = "/media/bit301/data/yml/data/p2_nii/external/lz/nor/PA185/2.nii.gz"
        path_save = path.replace("external", "Aortic_index")  # lo:97  x1:54  xlo2:10
        out_put = path_save.split("2.nii.gz")[0]
        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        aortic_index(path, path_save)
        pathh=path.replace("external", "test/MedNeXtx22")
        path_savee = path_save.replace("2.nii.gz", "22.nii.gz")
        aortic_index(pathh, path_savee)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

        # revision
        # path_gd = path_save.replace("2.nii.gz", "2.h5")
        # with h5py.File(path_gd, 'r') as f_gd:  # 评估四分位数据
        #     aa = f_gd['calcium_per_index'][:]
        #     aa = np.nan_to_num(aa)
        #     if np.all(aa == 1):
        #         aortic_index(path,path_save)
        #         print("I love China1!")
        #
        # path_mea = path_savee.replace("22.nii.gz", "22.h5")
        # with h5py.File(path_mea, 'r') as f_mea:  # 评估四分位数据
        #     aa = f_mea['calcium_per_index'][:]
        #     aa = np.nan_to_num(aa)
        #     if np.all(aa == 1):
        #         aortic_index(pathh, path_savee)
        #         print("I love China2!")

def compute_ccc(x, y,rho):
    '''
    计算一致性相关系数（Concordance Correlation Coefficient, CCC）
    '''
    # rho, _ = pearsonr(x, y)
    mean_x = np.mean(x)
    mean_y = np.mean(y)
    var_x = np.var(x)
    var_y = np.var(y)
    # 计算CCC系数
    ccc = 2 * rho * np.sqrt(var_x) * np.sqrt(var_y) / (var_x + var_y + (mean_x - mean_y) ** 2) + (mean_x - mean_y) ** 2 / (
                var_x + var_y + (mean_x - mean_y) ** 2)
    return ccc

def metric1(mea,gd):
    measurement_values = np.nan_to_num(mea)  # 将NaN替换为0，Inf替换为最大或最小浮点数
    groundtruth_values = np.nan_to_num(gd)
    mae = mean_absolute_error(groundtruth_values, measurement_values)# 计算MAE
    mre=my_mape(np.nanmean(groundtruth_values), np.nanmean(measurement_values))
    if mae==0:
        pearson_corr=1
        ccc=1
    else:
        # variance = np.var(measurement_values - groundtruth_values)# 计算方差
        # mape=mean_absolute_percentage_error(groundtruth_values, measurement_values)
        # r2=r2_score(measurement_values, groundtruth_values)
        if np.all(measurement_values[0] == measurement_values) and np.all(groundtruth_values[0] == groundtruth_values):
            pearson_corr=1
        else:
            if np.all(measurement_values[0] == measurement_values):
                measurement_values=np.where(groundtruth_values>0,1,0)*1e-8
            elif np.all(groundtruth_values[0] == groundtruth_values):
                groundtruth_values=np.where(measurement_values>0,1,0)*1e-8
            pearson_corr, _ = pearsonr(measurement_values, groundtruth_values)# 计算Pearson相关系数
        ccc = compute_ccc(groundtruth_values, measurement_values,pearson_corr)#不达预期

    # print("MAE: ", mae)
    # print("Pearson Correlation Coefficient: ", pearson_corr)
    # print("CCC: ", ccc)
    # print("\b")
    return mae,mre,pearson_corr,ccc

def metric(diameter_per_total,mea,gd):
    #The distal abdominal aorta (i.e., the portion near the bifurcation that divides the right and left iliac arteries)
    # has a diameter of approximately 1.5 cm to 2.0 cm (or 15 mm to 20 mm)
    threshold = 15 #approximate 10 mm
    above_threshold = diameter_per_total > threshold
    greater_indices = np.where(above_threshold)[0]
    if greater_indices.size > 0:
        start_index = greater_indices[0]
        end_index = greater_indices[-1] + 1  # 注意这里需要加1，因为end_index是要包含在内的

        mea=mea[start_index:end_index]
        gd=gd[start_index:end_index]
        measurement_values = np.nan_to_num(mea)  # 将NaN替换为0，Inf替换为最大或最小浮点数
        groundtruth_values = np.nan_to_num(gd)
        mae = mean_absolute_error(groundtruth_values, measurement_values)# 计算MAE
        # mapes,_=my_mape1(groundtruth_values, measurement_values)
        # mape=np.mean(mapes[start_index:end_index])
        mape=my_mape(np.nanmean(gd),np.nanmean(mea))
        if mae==0:
            pearson_corr=1
            ccc=1
        else:
            # variance = np.var(measurement_values - groundtruth_values)# 计算方差
            # mape=mean_absolute_percentage_error(groundtruth_values, measurement_values)
            # r2=r2_score(measurement_values, groundtruth_values)
            if np.all(measurement_values[0] == measurement_values) and np.all(groundtruth_values[0] == groundtruth_values):
                pearson_corr=1
            else:
                if np.all(measurement_values[0] == measurement_values):
                    measurement_values=np.where(groundtruth_values>0,1,0)*1e-8
                elif np.all(groundtruth_values[0] == groundtruth_values):
                    groundtruth_values=np.where(measurement_values>0,1,0)*1e-8
                pearson_corr, _ = pearsonr(measurement_values, groundtruth_values)# 计算Pearson相关系数
            ccc = compute_ccc(groundtruth_values, measurement_values,pearson_corr)#不达预期

        # print("MAE: ", mae)
        # print("Pearson Correlation Coefficient: ", pearson_corr)
        # print("CCC: ", ccc)
        # print("\b")
        return mae,mape,pearson_corr,ccc
    else:
        return 0,0,0,0

def calcium_alert_index(diameter_per_lumen, calcium_per_index_mea, calcium_per_index):
    # 指定一个阈值
    threshold = 15
    above_threshold = diameter_per_lumen > threshold
    greater_indices = np.where(above_threshold)[0]
    if greater_indices.size > 0:
        start_index = greater_indices[0]
        end_index = greater_indices[-1] + 1  # 注意这里需要加1，因为end_index是要包含在内的

        # 将原数组中不在指定区间的数值赋值为0
        calcium_per_index_mea[:start_index] = 0
        calcium_per_index_mea[end_index:] = 0
        cai_mea = np.percentile(calcium_per_index_mea, 75)
        sorted_arr = np.sort(calcium_per_index_mea)[::-1]
        top_mea = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])

        calcium_per_index[:start_index] = 0
        calcium_per_index[end_index:] = 0
        cai_gd = np.percentile(calcium_per_index, 75)
        sorted_arr = np.sort(calcium_per_index)[::-1]
        top_gd = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])
        return cai_mea, top_mea, cai_gd, top_gd

        # indices = np.nonzero(calcium_per_index_mea)
        # if len(indices[0])<2:
        #     cai_mea=0
        #     top_mea=0
        # else:
        #     calcium_per_index_mea = calcium_per_index_mea[indices]
        #     cai_mea = np.percentile(calcium_per_index_mea, 75)
        #     sorted_arr = np.sort(calcium_per_index_mea)[::-1]
        #     top_mea = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])


        # indices = np.nonzero(calcium_per_index)
        # if len(indices[0])<20:
        #     cai_gd=0
        #     top_gd=0
        # else:
        #     calcium_per_index = calcium_per_index[indices]
        #     cai_gd = np.percentile(calcium_per_index, 75)
        #     sorted_arr = np.sort(calcium_per_index)[::-1]
        #     top_gd = np.mean(sorted_arr[:int(len(sorted_arr) * 0.25)])
        # return cai_mea,top_mea, cai_gd,top_gd
    else:
        return 0,0,0,0

#ccc r^2
def statis():
    data = openpyxl.load_workbook('demo1.xlsx')
    table = data.active
    path = "/media/bit301/data/yml/data/p2_nii/Aortic_index/cq/"
    save_name=path.split("/")[-2]
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    ii = 1
    lumenn=[]
    totall=[]
    calciumm=[]
    flag_gd=[]
    flag_mea=[]
    for path in path_list:
        calcium_gd = 0
        calcium_mea = 0
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_lumen = f_gd['diameter_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            calcium_per_index = f_gd['calcium_per_index'][:]
            total_calcium_index = f_gd['total_calcium_index'][:]

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
            diameter_per_total_mea = f_mea['diameter_per_total'][:]
            calcium_per_index_mea = f_mea['calcium_per_index'][:]
            total_calcium_index_mea = f_mea['total_calcium_index'][:]
        if np.all(calcium_per_index == 1) or np.all(calcium_per_index_mea == 1):
            print(path+"_gd")
        elif np.all(diameter_per_total == 0):
            print(path+"_mea")
        lumnen = metric(diameter_per_total,diameter_per_lumen_mea, diameter_per_lumen)
        total = metric(diameter_per_total,diameter_per_total_mea, diameter_per_total)
        calcium = metric(diameter_per_total,calcium_per_index_mea, calcium_per_index)
        lumenn.append(lumnen)
        totall.append(total)
        calciumm.append(calcium)

        write_path = path.split('Aortic_index')[1]
        table.cell(ii + 1, 3).value = write_path
        # if total_calcium_index[1]>0.075:
        #     calcium_gd=1
        # if total_calcium_index_mea[1] > 0.065:
        #     calcium_mea=1

        cai_mea,top_mea, cai_gd,top_gd=calcium_alert_index(diameter_per_lumen,calcium_per_index_mea, calcium_per_index)
        if cai_gd>=0.075 or top_gd>=0.2:
            calcium_gd=1
        if cai_mea >= 0.07 or top_mea >= 0.18:
            calcium_mea=1
        # if top_gd>0.2:
        #     calcium_gd=1
        # if top_mea > 0.18:
        #     calcium_mea=1

        # if cai_gd>0.075:
        #     calcium_gd=1
        # if cai_mea > 0.065:
        #     calcium_mea=1
        table.cell(ii + 1, 4).value = calcium_gd#标注钙化提示
        table.cell(ii + 1, 5).value = calcium_mea #标注钙化提示
        flag_gd.append(calcium_gd)
        flag_mea.append(calcium_mea)

        # if ii==38:
        #     a=1
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    data.save(save_name+'_metricx.xlsx') #
    mean_lumnen = np.nanmean(lumenn, axis=0)  # 列平均
    print("mean_lumnen:", mean_lumnen)
    mean_total = np.nanmean(totall, axis=0)  # 列平均
    print("mean_vaseel:", mean_total)
    mean_calcium = np.nanmean(calciumm, axis=0)  # 列平均
    print("mean_calcium:", mean_calcium)
    cm = confusion_matrix(flag_gd, flag_mea)
    print("confusion_matrix:", cm)
    print("finished!")

#管腔，血管，钙化指数 曲线
def disp1():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/disp.txt")  # hnnk test.txt
    path_list = []
    for line in f.readlines():#tile_step_size=0.75较好处理官腔错位问题
            path=line.split('\n')[0]
            path_list.append(path)
    ij=0
    for path in path_list:
        calcium_gd = 0
        calcium_mea = 0
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_lumen = f_gd['diameter_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            calcium_per_index = f_gd['calcium_per_index'][:]
            total_calcium_index = f_gd['total_calcium_index'][:]

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
            diameter_per_total_mea = f_mea['diameter_per_total'][:]
            calcium_per_index_mea = f_mea['calcium_per_index'][:]
            total_calcium_index_mea = f_mea['total_calcium_index'][:]

        max_len = len(diameter_per_lumen)
        fig, axs = plt.subplots(nrows=1, ncols=3, figsize=(5, 10), sharey=True, layout='constrained')#(4, 12)
        line_color1 = 'b'
        line_color2 = 'r'
        line_width = 1
        # q1_color = "deepskyblue"
        # q2_color = "r"
        for ax in axs:
            ax.invert_yaxis()

        axs[0].plot(diameter_per_lumen[::-1], range(max_len), color=line_color1, linewidth=line_width, label='True')
        axs[0].plot(diameter_per_lumen_mea[::-1], range(max_len), color=line_color2, linewidth=line_width, label='Pred')
        axs[0].set_title('Lumen',fontsize=10)
        axs[0].set_xlabel('Diameter', fontsize=10)  # 添加横轴标题 pixel
        axs[0].set_ylabel('Slice Number', fontsize=10)  # 添加纵轴标题
        axs[0].legend(loc='best', fontsize=10)  # 添加图例，loc='best' 表示自动寻找最佳位置放置图例

        axs[1].plot(diameter_per_total[::-1], range(max_len), color=line_color1, linewidth=line_width, label='True')
        axs[1].plot(diameter_per_total_mea[::-1], range(max_len), color=line_color2, linewidth=line_width, label='Pred')
        axs[1].set_title('Vessel ',fontsize=10)
        axs[1].set_xlabel('Diameter', fontsize=10)  # 添加横轴标题
        axs[1].legend(loc='best', fontsize=10)  # 添加图例，loc='best' 表示自动寻找最佳位置放置图例

        axs[2].plot(calcium_per_index[::-1], range(max_len), color=line_color1, linewidth=line_width, label='True')
        axs[2].plot(calcium_per_index_mea[::-1], range(max_len), color=line_color2, linewidth=line_width, label='Pred')
        axs[2].set_title('Calcification',fontsize=10)
        axs[2].set_xlabel('Index', fontsize=10)  # 添加横轴标题
        axs[2].legend(loc='best', fontsize=10)  # 添加图例，loc='best' 表示自动寻找最佳位置放置图例
        # plt.savefig("High resoltion.png", dpi=600)
        out_put = "disp"
        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        file=path.split("Aortic_index/")[1].split("/2")[0].replace("/","_")+".tif"
        save_path = os.path.join(out_put, file)
        plt.savefig(save_path)#矢量图
        # plt.savefig(save_path, dpi=600)
        # plt.show()
        # 显式关闭当前figure
        plt.close(fig)

def per_mean1(diameter_per_lumen, calcium_per_index,calcium_per_index_mea):
    arr1 = calcium_per_index
    arr2 = calcium_per_index_mea
    merged_indices = np.nonzero(arr1 + arr2)
    if len(merged_indices):
        non_zero_values1 = arr1[merged_indices]
        m1 = np.mean(non_zero_values1)  # 像钙化指数  groundtruth或reconstruction存在0/0，则会出现警告
        non_zero_values2 = arr2[merged_indices]
        m2 = np.mean(non_zero_values2)
        return m1, m2
    else:
        return 0, 0

def per_mean(diameter_per_lumen, calcium_per_index, calcium_per_index_mea):
    threshold =15
    above_threshold = diameter_per_lumen > threshold
    greater_indices = np.where(above_threshold)[0]
    if greater_indices.size > 0:
        start_index = greater_indices[0]
        end_index = greater_indices[-1] + 1  # 注意这里需要加1，因为end_index是要包含在内的
        calcium_per_index_mea[:start_index] = 0
        calcium_per_index_mea[end_index:] = 0
        calcium_per_index[:start_index] = 0
        calcium_per_index[end_index:] = 0
        arr1=calcium_per_index
        arr2=calcium_per_index_mea
        # mapes,merged_indices=my_mape1(arr1, arr2)
        # mape = np.mean(mapes[start_index:end_index])
        m1 = np.mean(arr2[start_index:end_index])  # 像钙化指数  groundtruth或reconstruction存在0/0，则会出现警告
        m2 = np.mean(arr1[start_index:end_index])
        mape=my_mape(m2, m1)
        return m1, m2,mape
    else:
        return 0, 0


def dot_plot(data,save_path):
    gd="True "+save_path.split("_")[-1]
    mea="Predict "+save_path.split("_")[-1]+" from NCCT"
    save_path=save_path+".tif"
    data=np.array(data)
    # true_values = data[:, 0]  # 请替换为实际真实值数组
    # predicted_values = data[:, 1]
    # errors = np.abs(predicted_values - true_values)
    true_values = data[:, 0]  # 请替换为实际真实值数组
    predicted_values = data[:, 1]
    errors = data[:, 2]
    fig, ax = plt.subplots()
    scatter = ax.scatter(true_values, predicted_values, c=errors, cmap='viridis', s=20, alpha=0.8)
    cbar = fig.colorbar(scatter, ax=ax, label='Mean Absolute Percentage Error (%)')
    plt.plot([np.nanmin(true_values), np.nanmax(true_values)],
             [np.nanmin(true_values), np.nanmax(true_values)],
             'r--', label='Perfect reconstruction line')
    ax.set_xlabel(gd, fontsize=10)
    ax.set_ylabel(mea, fontsize=10)
    plt.legend()
    ax.set_title('Predict vs True with Mean Absolute Percentage Error', fontsize=10)
    # plt.savefig(save_path)  # 矢量图
    plt.savefig(save_path, dpi=600)
    # plt.show()
    plt.close(fig)# 显式关闭当前figure

#Bland-Altman，
def disp2():
    path = "/media/bit301/data/yml/data/p2_nii/Aortic_index/cq/"#
    save_name=path.split("/")[-2]
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "/2.h5" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    lumen_total = []
    vessel_total = []
    calcium_total = []
    for path in path_list:
        with h5py.File(path, 'r') as f_gd:#评估四分位数据
            diameter_per_lumen = f_gd['diameter_per_lumen'][:]
            diameter_per_total = f_gd['diameter_per_total'][:]
            calcium_per_index = f_gd['calcium_per_index'][:]
            total_calcium_index = f_gd['total_calcium_index'][:]

        path_mea=path.replace("2.h5","22.h5")
        with h5py.File(path_mea, 'r') as f_mea:#评估四分位数据
            diameter_per_lumen_mea = f_mea['diameter_per_lumen'][:]
            diameter_per_total_mea = f_mea['diameter_per_total'][:]
            calcium_per_index_mea = f_mea['calcium_per_index'][:]
            total_calcium_index_mea = f_mea['total_calcium_index'][:]
        if np.all(calcium_per_index == 1) or np.all(calcium_per_index_mea == 1):
            print(path+"_gd")
        elif np.all(diameter_per_total == 0):
            print(path+"_mea")
        lumnen = per_mean(diameter_per_lumen,diameter_per_lumen,diameter_per_lumen_mea)
        total = per_mean(diameter_per_lumen,diameter_per_total,diameter_per_total_mea)
        calcium = per_mean(diameter_per_lumen,calcium_per_index,calcium_per_index_mea)
        lumen_total.append(lumnen)
        vessel_total.append(total)
        calcium_total.append(calcium)
    out_put = "scatter_disp"
    if not os.path.isdir(out_put):
        os.makedirs(out_put)
    save_path = os.path.join(out_put, save_name+"_lumen diameter")
    dot_plot(lumen_total, save_path)
    save_path = os.path.join(out_put, save_name+"_vessel diameter")
    dot_plot(vessel_total, save_path)
    save_path = os.path.join(out_put, save_name+"_calcium index")
    dot_plot(calcium_total, save_path)

if __name__ == '__main__':
    # Aortic_index_caculate()
    # statis()
    # disp1()
    disp2()
