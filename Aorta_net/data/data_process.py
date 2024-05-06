import numpy as np
import matplotlib.pyplot as plt
import pydicom
import shutil
import random
import torch.nn as nn
import os
import SimpleITK as sitk
import ants
import cv2
import scipy
from skimage import measure
import pandas as pd

import nibabel as nib
import imageio
import itk
from PIL import Image
from scipy.stats import beta
import math
from natsort import natsorted
import copy


############# 文件操作模块 #############
def copy_and_paste():
    # path = "../../../data/diag_data/"  # CT_CTA disease
    catch = "../../../data/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)

    f = open("p_test1.txt")  # test train！！  list  filter
    path_list = []
    for line in f.readlines():
        if "SE0" in line and line.split('/IM')[0] not in path_list:
            path_list.append(line.split('/IM')[0])
    path_list.sort()
    ii = 0
    for sub_path in path_list:
        input_files = os.listdir(sub_path)
        input_files.sort()
        input_files.sort(key=lambda x: (int(x.split('IM')[1])))
        sub_pathe2 = sub_path.replace("SE0", "SE1")
        diag = sub_path.split("/")[4]
        sub_path_out1 = sub_path.replace(diag, "diag1a_data")
        sub_path_out2 = sub_path_out1.replace("SE0", "SE1")
        if not os.path.isdir(sub_path_out1):
            os.makedirs(sub_path_out1)

        if not os.path.isdir(sub_path_out2):
            os.makedirs(sub_path_out2)

        for j in range(len(input_files)):
            in_file_path1 = os.path.join(sub_path, input_files[j])
            in_file_path2 = os.path.join(sub_pathe2, input_files[j])
            out_file_path1 = os.path.join(sub_path_out1, input_files[j])
            out_file_path2 = os.path.join(sub_path_out2, input_files[j])
            shutil.copy(in_file_path1, out_file_path1)
            shutil.copy(in_file_path2, out_file_path2)

        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)


# 将各个子文件夹合并到一起
def mv_file():
    # path = "../../../data/diag_data/"  # CT_CTA disease
    catch = "../../../data/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)

    input = '../output/Cyc/1e34/'  ######
    output = '../output/diag1a_data/'  ######
    # output = '/media/yml/yml/data/make_choice/diag0a_data/'  ######
    if not os.path.isdir(input):
        os.makedirs(input)
    if not os.path.isdir(output):
        os.makedirs(output)
    path_list = []
    for root, dirs, files in os.walk(input, topdown=False):
        if "SE1" in root:
            path_list.append(root)
    path_list.sort()

    ii = 0
    for sub_path in path_list:
        input_files = os.listdir(sub_path)
        input_files.sort()
        input_files.sort(key=lambda x: (int(x.split('IM')[1])))
        sub_out = sub_path.replace(input, output)
        sub_out = sub_out.replace("SE1", "SE3")  #######
        if not os.path.isdir(sub_out):
            os.makedirs(sub_out)

        for j in range(len(input_files)):
            in_file_path = os.path.join(sub_path, input_files[j])
            out_file_path = os.path.join(sub_out, input_files[j])
            shutil.move(in_file_path, out_file_path)

        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)


def remove_file():
    path = "../../../data/p2_nii/"  # CT_CTA disease
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        if "4.nii.gz" in files:
            path = os.path.join(root, '4.nii.gz')
            # path_list.append(path)
            os.remove(path)
    # path_list.sort()
    # for sub_path in path_list:
    #     aa = os.path.join(sub_path.split('SE0')[0], 'SE2')
    #     if os.path.isdir(aa):
    #         shutil.rmtree(aa)


# get real input list
def get_test_list():
    path = "../../../data/p2_nii/"  # neckk
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        path_list.append(root)
    f = open("nor.txt", "w")
    for sub_path in path_list:
        data_files = os.listdir(sub_path)
        for j in range(len(data_files)):
            file_path = os.path.join(sub_path, data_files[j])
            if os.path.isdir(file_path):
                continue
            # file_path_list.append(file_path)
            f.writelines(file_path + "\n")
    f.close()


# 主函数
def nii_to_image():
    filepath = "../../../data/ASL/"
    imgfile = "../../../data/ASL_img/"
    filenames = os.listdir(filepath)

    # 开始读取nii文件
    for f in filenames:
        img_path = os.path.join(filepath, f)
        img = nib.load(img_path)  # 读取nii
        img_fdata = img.get_fdata()
        # 去掉nii的后缀名
        fname = f.replace('.nii', '')
        img_f_path = os.path.join(imgfile, fname)

        # 创建nii对应的图像的文件夹
        if not os.path.isdir(img_f_path):
            os.makedirs(img_f_path)
        # if not os.path.exists(img_f_path):
        #     # 新建文件夹
        #     os.mkdir(img_f_path)
        #     # 开始转换为图像
        (x, y, z) = img.shape
        # z是图像的序列
        for i in range(z):
            # 选择哪个方向的切片都可以
            silce = img_fdata[:, :, i]
            # silce = (silce * 255.0).astype('uint8')
            # silce = silce.astype(np.uint8)
            imageio.imwrite(os.path.join(img_f_path, '{}.png'.format(i)), silce)

            # silce = silce.astype(np.uint8)
            # # silce = (silce * 255.0).astype('uint8')
            # silce = Image.fromarray(silce)
            # silce.save(os.path.join(img_f_path, '{}.png'.format(i)))


############# 自动构建mask #############
def to_windowdata(image, WC, WW):
    # image = (image + 1) * 0.5 * 4095
    # image[image == 0] = -2000
    # image=image-1024
    center = WC  # 40 400//60 300
    width = WW  # 200
    try:
        win_min = (2 * center - width) / 2.0 + 0.5
        win_max = (2 * center + width) / 2.0 + 0.5
    except:
        # print(WC[0])
        # print(WW[0])
        center = WC[0]  # 40 400//60 300
        width = WW[0]  # 200
        win_min = (2 * center - width) / 2.0 + 0.5
        win_max = (2 * center + width) / 2.0 + 0.5
    dFactor = 255.0 / (win_max - win_min)
    image = image - win_min
    image = np.trunc(image * dFactor)
    image[image > 255] = 255
    image[image < 0] = 0
    image = image / 255  # np.uint8(image)
    # image = (image - 0.5)/0.5
    return image


def remove_small_points(img, threshold_point=20):  # 80
    img_label, num = measure.label(img, return_num=True,connectivity=2)  # 输出二值图像中所有的连通域
    props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等
    resMatrix = np.zeros(img_label.shape)
    dia = 6  # 60
    for i in range(1, len(props)):
        are = props[i].area
        if are > threshold_point:
            if are < 3400:
                dia = int(np.sqrt(are) / 2)
                tmp = (img_label == i + 1).astype(np.uint8)
                kernel = np.ones((dia, dia), np.uint8)
                tmp = cv2.dilate(tmp, kernel)  # [-1 1]
                # x, y = np.nonzero(tmp)
                # tmp[x[0] - dia:x[-1] + dia, y[0] - dia:y[-1] + dia] = 1  # resize
                resMatrix += tmp  # 组合所有符合条件的连通域
    resMatrix *= 1
    resMatrix[resMatrix > 1] = 1  #
    return resMatrix

def vascular_mask(file_path1, file_path2):
    # file_pathE = '../output/Hd/img2/e1/4/ST0/SE0/IM11'
    # file_path1 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE0/IM6'

    dicom = sitk.ReadImage(file_path1)
    data1 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data1 = data1 + 1024
    # file_path2 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE1/IM6'
    dicom = sitk.ReadImage(file_path2)
    data2 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data2 = data2 + 1024

    image1 = data1  # sitk读取的数值比pydicom读取的数值小1024
    image1[image1 < 1037] = -2000  # -2000->0
    image1[image1 > 1100] = -2000
    image1 = image1 / 4095
    image1 = (image1 - 0.5) / 0.5
    image1[image1 > -1] = 1
    image1[image1 <= -1] = 0

    image2 = data2  # sitk读取的数值比pydicom读取的数值小1024
    image2 = to_windowdata(image2, 120, 20)
    image2 = image2 * image1
    image2[image2 <= 0] = -1
    kernel1 = np.ones((2, 2), np.uint8)
    image2 = cv2.dilate(image2, kernel1)  # [-1 1]
    image2 = remove_small_points(image2, 30)
    return image2


def make_mask():
    path = "../../../data/p1_data/"  # CT_CTA disease
    catch = "../../../data/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)

    # path_list=[]
    # for root, dirs, files in os.walk(path, topdown=False):
    #     if"SE0" in root:
    #         path_list.append(root)
    # path_list.sort()

    f = open("p1_data.txt")  # test train！！  list  filter
    path_list = []
    for line in f.readlines():
        if "SE0" in line and line.split('/IM')[0] not in path_list:
            path_list.append(line.split('/IM')[0])
    path_list.sort()
    ii = 0
    for sub_path in path_list:
        input_files = os.listdir(sub_path)
        input_files.sort()
        input_files.sort(key=lambda x: (int(x.split('IM')[1])))
        putpath2 = sub_path.replace("SE0", "SE2")
        if os.path.isdir(putpath2):
            shutil.rmtree(putpath2)

        if not os.path.isdir(putpath2):
            os.makedirs(putpath2)
        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(input_files)):
            try:
                file_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
                # file_path='../../../data/p1_data/PA0/SE0/IM1'
                # dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
                dsB = pydicom.dcmread(file_path.replace("SE0", "SE1"), force=True)  # 读取头文件
                mask = vascular_mask(file_path, file_path.replace("SE0", "SE1"))
                mask[mask <= 0.1] = -2000
                mask[mask > 0.1] = 2095
                dsB.SeriesInstanceUID = dsbsi
                dsB.SeriesNumber = 5
                newimgb = mask.astype(np.int16)
                dsB.PixelData = newimgb.tobytes()
                out_path = os.path.join(file_path.replace("SE0", "SE2"))
                pydicom.dcmwrite(out_path, dsB)
            except:
                continue
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    return 0


def display():  # file_path1,file_path2
    # file_pathE = '../output/Hd/img2/e1/4/ST0/SE0/IM11'
    file_path1 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE0/IM6'

    dicom = sitk.ReadImage(file_path1)
    data1 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data1 = data1 + 1024
    file_path2 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE1/IM6'
    dicom = sitk.ReadImage(file_path2)
    data2 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data2 = data2 + 1024
    data = data2 - data1
    dsA = pydicom.dcmread(file_path1, force=True)
    # aa=ds.SeriesInstanceUID
    # ab=ds.FrameOfReferenceUID
    # ac=ds.StudyInstanceUID
    #
    # file_pathE = '../output/Hd1/img2/e40/4/ST0/SE1/IM11'
    # dsb = pydicom.dcmread(file_pathE, force=True)  # 读取头文件
    # aa2=dsb.SeriesInstanceUID
    # ab2=dsb.FrameOfReferenceUID
    # ac2=dsb.StudyInstanceUID

    # data = (ds.pixel_array).astype(np.int)
    # WC = ds.WindowCenter
    # WW = ds.WindowWidth
    # a = to_windowdata(data, WC, WW)
    image1 = data1  # sitk读取的数值比pydicom读取的数值小1024
    image1 = image1 / 4095
    image1 = (image1 - 0.5) / 0.5
    image1[image1 > 1] = 1
    image1[image1 <= -1] = -1
    # kernel2 = np.ones((2, 2), np.uint8)
    # image1= cv2.dilate(image1, kernel2)  # [-1 1]

    image2 = data1  # sitk读取的数值比pydicom读取的数值小1024
    image2[image2 < 1037] = -2000  # -2000->0
    image2[image2 > 1100] = -2000
    image2 = image2 / 4095
    image2 = (image2 - 0.5) / 0.5
    image2[image2 > -1] = 1
    image2[image2 <= -1] = 0
    # kernel2 = np.ones((2, 0), np.uint8)
    # image2 = cv2.dilate(image2, kernel2)#[-1 1]

    image3 = data2  # sitk读取的数值比pydicom读取的数值小1024
    image3[image3 < 0] = 0  # -2000->0
    image3 = image3 / 4095
    image3 = (image3 - 0.5) / 0.5

    image4 = data2  # sitk读取的数值比pydicom读取的数值小1024
    image4 = to_windowdata(image4, 120, 20)
    image5 = image4 * image2
    image5[image5 <= 0] = -1
    kernel1 = np.ones((2, 2), np.uint8)
    image5 = cv2.dilate(image5, kernel1)  # [-1 1]
    image5 = remove_small_points(image5, 30)
    # m=get_mask(image2)
    # image2=image2*m
    # image2[image2==0]=-1

    # image3=image1*image2
    # image3[image3 == 0] = -1

    # plt.close()
    plt.subplot(2, 2, 1)
    plt.imshow(image1, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 2)
    plt.imshow(image2, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 3)
    plt.imshow(image4, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 4)
    plt.imshow(image5, cmap='gray')  # ,vmin=0,vmax=255
    plt.show()
    # cv2.imwrite("outpath"+'.png',((image2*0.5)+0.5)*255)


def resample(image, scan, new_spacing=[1.25, 0.67, 0.67]):
    # Determine current pixel spacing
    # [1.25,0.67,0.67]
    try:
        spacing = np.array([scan.SliceThickness, scan.PixelSpacing[0], scan.PixelSpacing[1]], dtype=np.float32)
    except:
        spacing = np.array([scan.SliceThickness, scan.PixelSpacing, scan.PixelSpacing], dtype=np.float32)
    # new_spacing=np.array([z_spacing,scan.PixelSpacing[0], scan.PixelSpacing[1]], dtype=np.float32)
    resize_factor = spacing / new_spacing
    new_real_shape = image.shape * resize_factor
    new_shape = np.round(new_real_shape)
    real_resize_factor = new_shape / image.shape
    new_spacing = spacing / real_resize_factor
    # image = scipy.ndimage.interpolation.zoom(image, real_resize_factor, mode='nearest').transpose((1,2,0)) #'nearest'  'bilinear'
    image = scipy.ndimage.zoom(image, real_resize_factor, mode='nearest').transpose((1, 2, 0))
    image = padding_and_cropping(image, (512, 512))
    # image=resize_image_with_crop_or_pad(image,(512,512))
    return image, new_spacing,


# slicers排序+重采样+get list（DICOM文件）301
def make_process():
    path = "../../../../data/p3/jfj/"  # CT_CTA disease jfj3/xz/DICOM1
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        if "SE0" in root:
            path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)

    i = 0
    id = []
    for sub_path in path_list:
        # ii=ii+1
        # if ii<3:
        #     continue
        input_files = os.listdir(sub_path)
        input_files.sort()  #
        le = len(input_files)
        if le < 200 or le > 2000:
            shutil.rmtree(sub_path)
            continue

        # get file list
        DCMA = []
        loca = np.zeros((1, len(input_files)))
        file_path = os.path.join(sub_path, input_files[0])  # 直接将文件中按行读到list里，效果与方法2一样
        dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
        if dsA.PatientID not in id:
            id.append(dsA.PatientID)
        else:
            print("ID 重复")
            print(sub_path)
            shutil.rmtree(sub_path.split("/ST")[0])
            continue

        for j in range(len(input_files)):
            file_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
            DCMA.append(dsA)
            loca[0, j] = dsA.SliceLocation  ##层面是否对应

        # 对slicers根据location进行排序
        file_n = len(np.unique(loca))
        idxa = np.argsort((-loca))
        imga = np.zeros([file_n, 512, 512])
        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n):  # 排序，保证命名和序列一致

            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsA.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        imga[imga == 0] = -2000
        outa_dicom, new_spacing = resample(imga, dsA)

        # save to dicom file
        putpath1 = sub_path.split("/SE")[0]
        shutil.rmtree(sub_path.split("/ST")[0])  # 写入重采样文件前先删除原始文件
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
        dsbsi = pydicom.uid.generate_uid()
        for k in range(outa_dicom.shape[2]):
            dssopc = pydicom.uid.generate_uid()
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1
            dsA.SOPInstanceUID = dssopc  # needed for 3D slicer
            dsA.SeriesDescription = 'CTA 1.25mm'

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)

        se0output = os.path.join(sub_path.split("/ST")[0], '0.nii.gz')
        dcm2nii_sitk(putpath1, se0output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)
    print('finished!')


# xy
def make_process_xy():
    path = "../../../../data/p3/xy/"  # CT_CTA disease jfj3/xz/DICOM1

    # k = 0
    # for root, dirs, files in os.walk(path, topdown=False):
    #     if"CTA" in root and "PA" not in root:
    #         oldfile_name=root.split("/CTA")[0]
    #         file_name=root.split("/")[-2]
    #         newname=oldfile_name.replace(file_name,"PA"+str(k))
    #         os.rename(oldfile_name, newname)
    #         k=k+1

    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        if "CTA" in root:
            path_list.append(root)
        elif "Body" in root:
            shutil.rmtree(root)
    path_list.sort()
    # random.shuffle(path_list)

    i = 0
    id = []
    for sub_path in path_list:
        # ii=ii+1
        # if ii<3:
        #     continue
        input_files = os.listdir(sub_path)
        input_files.sort()  #
        le = len(input_files)
        if le < 200 or le > 3000:
            shutil.rmtree(sub_path)
            continue

        # get file list
        DCMA = []
        loca = np.zeros((1, len(input_files)))
        file_path = os.path.join(sub_path, input_files[0])  # 直接将文件中按行读到list里，效果与方法2一样
        dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
        # if dsA.PatientID not in id:
        #     id.append(dsA.PatientID)
        # else:
        #     print("ID 重复")
        #     print(sub_path)
        #     shutil.rmtree(sub_path.split("/CTA")[0])
        #     continue

        for j in range(len(input_files)):
            file_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
            DCMA.append(dsA)
            loca[0, j] = dsA.SliceLocation  ##层面是否对应

        # 对slicers根据location进行排序
        file_n = len(np.unique(loca))
        idxa = np.argsort((-loca))
        imga = np.zeros([file_n, 512, 512])
        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n):  # 排序，保证命名和序列一致

            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsA.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        # imga[imga==0]=-2000
        outa_dicom, new_spacing = resample(imga, dsA)

        # save to dicom file
        putpath1 = sub_path  # .split("/SE")[0]
        shutil.rmtree(sub_path.split("/CTA")[0])  # 写入重采样文件前先删除原始文件
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
        dsbsi = pydicom.uid.generate_uid()
        for k in range(outa_dicom.shape[2]):
            dssopc = pydicom.uid.generate_uid()
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1
            dsA.SOPInstanceUID = dssopc  # needed for 3D slicer
            dsA.SeriesDescription = 'CTA 1.25mm'

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)

        se0output = os.path.join(sub_path.split("/CTA")[0], '0.nii.gz')
        dcm2nii_sitk(putpath1, se0output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)
    print('finished!')


# hnnk   需要确定是保存为重采样格式还是整理格式
def make_abd_process1():
    path = "/media/yml/yml_data2/data_raw/hnnk/"  # neckk
    # path="../../../data/pv_data/jc/DICOM1/"
    # output = "../../../data/pv_data/hnnk/"
    output = "/media/yml/yml_data2/data_processed/FB/hnnk/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)

    f1 = open("hnnk.txt", "w")  # 564
    ij = 0
    id = []
    name = []
    for sub_path in path_list:
        # if ij<131:
        #     ij=ij+1
        #     continue
        data_files = os.listdir(sub_path)
        if len(data_files) < 350:
            continue
        putpath = os.path.join(output, sub_path.split("hnnk/")[1])
        file = {}
        # flag_s=0
        for i in range(len(data_files)):
            file_path = os.path.join(sub_path, data_files[i])
            ds = pydicom.dcmread(file_path, force=True)
            flag_s = str(ds.SeriesNumber)
            if flag_s not in file.keys():
                file[flag_s] = []
                file[flag_s].append(file_path)
            else:
                file[flag_s].append(file_path)
        x = list(file.keys())
        try:
            input_files = file[x[0]]
            target_files = file[x[1]]
        except:
            continue
        # if len(target_files)%len(input_files)!=0:
        #     continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        for j in range(len(input_files)):
            input_path = input_files[j]  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
            dsA.SeriesDescription = 'CT'
            dsA.SeriesNumber = 2
            DCMA.append(dsA)
            loca[0, j] = dsA.SliceLocation  ###层面是否对应

        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(target_files)):
            target_path = target_files[j]
            dsB = pydicom.dcmread(target_path, force=True)  #
            dsB.SeriesDescription = 'CTA'
            dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
            dsB.InstanceNumber = dsA.InstanceNumber
            dsB.SeriesInstanceUID = dsbsi  # 区分
            dsB.SeriesNumber = 3  # 区分
            DCMB.append(dsB)
            locb[0, j] = dsB.SliceLocation

        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].SliceLocation
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].SliceLocation:
                sl = DCMB[b].SliceLocation
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号

        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        if outb_dicom.shape[2] % outa_dicom.shape[2] != 0:
            continue
        elif outa_dicom.shape[2] < 200:
            continue

        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)

        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        if "neck" in sub_path:
            be = outa_dicom.shape[2] - 35
        else:
            be = 0
            num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1

            # dsB.SeriesNumber = 3  # 区分
            dsB.SliceLocation -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.SliceLocation]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
            f1.writelines(file_patha + "\n")
        try:
            if float(dsA.AcquisitionTime) > float(dsB.AcquisitionTime):
                os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
                os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
                os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        except:
            if float(dsA.InstanceCreationTime) > float(dsB.InstanceCreationTime):
                os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
                os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
                os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
            print(sub_path)
        ij = ij + 1
        # if ij % 10 == 0:
        print('numbers:', ij)
    # excel_name = "/p1_data_neck1_all.xlsx"
    # excel_path = os.path.dirname(os.path.abspath(__file__)) + excel_name
    # df = pd.DataFrame(columns=["ID","path", "date", "thickness/mm", "normal", "aneurysm", "stenosis"])
    # df.to_excel(excel_path, index=False)
    # df = pd.read_excel(excel_path, header=None)
    # ds = pd.DataFrame(name)
    # df = df.append(ds, ignore_index=True)
    # df.to_excel(excel_path, index=False, header=False)
    # f1.close()
    print('id numbers:', len(id))


def make_abd_process1b():
    path = "/media/yml/yml_data2/hnnk1/"  # neckk
    # path="../../../data/pv_data/jc/DICOM1/"
    # output = "../../../data/pv_data/hnnk/"
    output = "/media/yml/yml_data1/data/FB/hnnk/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)

    f1 = open("hnnk.txt", "w")  # 564
    ij = 0
    id = []
    name = []
    dsbsi = pydicom.uid.generate_uid()
    for sub_path in path_list:
        # if ij<131:
        #     ij=ij+1
        #     continue
        data_files = os.listdir(sub_path)
        if len(data_files) < 350:
            continue
        putpath = os.path.join(output, sub_path.split("hnnk1/")[1])
        file = {}
        # flag_s=0
        for i in range(len(data_files)):
            file_path = os.path.join(sub_path, data_files[i])
            ds = pydicom.dcmread(file_path, force=True)
            flag_s = str(ds.SeriesNumber)
            if flag_s not in file.keys():
                file[flag_s] = []
                file[flag_s].append(file_path)
            else:
                file[flag_s].append(file_path)
        x = list(file.keys())
        try:
            input_files = file[x[0]]
            target_files = file[x[1]]
        except:
            continue
        if len(input_files) != len(target_files):
            continue
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(input_files)))
        try:
            for j in range(len(input_files)):
                input_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
                target_path = os.path.join(sub_path, target_files[j])
                dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
                dsB = pydicom.dcmread(target_path,
                                      force=True)  # 读取头文件  1.2.840.113619.2.278.3.2831212038.186.1561204307.645
                if dsA.PatientID not in id:
                    id.append(dsA.PatientID)
                elif id[-1] != dsA.PatientID:  # 一个文件夹下应该为同一个人，如果出现不同的patient，则删除。
                    shutil.rmtree(sub_path)
                    print(sub_path)
                    continue
                if dsA.RescaleIntercept != -1024 or dsB.RescaleIntercept != -1024:  # check
                    print("dsA.RescaleIntercept!=-1024:", input_path)
                    break
                dsA.SeriesDescription = 'CT'
                dsA.SeriesNumber = 2

                dsB.SeriesDescription = 'CTA'
                dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
                dsB.InstanceNumber = dsA.InstanceNumber
                dsB.SeriesInstanceUID = dsbsi  # 区分
                dsB.SeriesNumber = 3  # 区分
                DCMA.append(dsA)
                DCMB.append(dsB)
                loca[0, j] = dsA.SliceLocation  ###层面是否对应
                locb[0, j] = dsB.SliceLocation
        except:
            print(sub_path)
            continue

        idxa = np.argsort((loca))
        idxb = np.argsort((locb))
        # putpath1 = putpath + str(i) + '/SE0'
        # putpath2 = putpath + str(i) + '/SE1'
        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        try:
            for k in range(len(input_files)):  # 排序，保证命名和序列一致
                if not os.path.isdir(putpath1):
                    os.makedirs(putpath1)
                    os.makedirs(putpath2)
                file_patha = os.path.join(putpath1, str(k + 1) + '.dcm')
                file_pathb = os.path.join(putpath2, str(k + 1) + '.dcm')
                a = idxa[0, k]
                b = idxb[0, k]
                pydicom.dcmwrite(file_patha, DCMA[a])
                pydicom.dcmwrite(file_pathb, DCMB[b])
        except:
            print(sub_path)
            continue
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    # excel_name = "/p1_data_neck1_all.xlsx"
    # excel_path = os.path.dirname(os.path.abspath(__file__)) + excel_name
    # df = pd.DataFrame(columns=["ID","path", "date", "thickness/mm", "normal", "aneurysm", "stenosis"])
    # df.to_excel(excel_path, index=False)
    # df = pd.read_excel(excel_path, header=None)
    # ds = pd.DataFrame(name)
    # df = df.append(ds, ignore_index=True)
    # df.to_excel(excel_path, index=False, header=False)
    # f1.close()
    print('id numbers:', len(id))


# xy
def make_abd_process2():
    path = "/media/yml/yml_data2/data_raw/xy3/"  # neckk
    # path="../../../data/pv_data/jc/DICOM1/"
    output = "/media/yml/yml_data2/data_processed/FB/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)

    f1 = open("xy3.txt", "w")  # 564
    ij = 1
    id = []
    name = []
    for sub_path in path_list:
        # if ij<25:
        #     ij=ij+1
        #     continue
        putpath = os.path.join(output, "PA" + str(ij))
        file_list = os.listdir(sub_path)
        # ncct=[]
        # cta=[]
        for file in file_list:
            if "CTA" not in file:
                subb_path = os.path.join(sub_path, file)
                dicom_list = os.listdir(subb_path)
                input_files = [os.path.join(subb_path, x) for x in dicom_list]

            else:
                subb_path = os.path.join(sub_path, file)
                dicom_list = os.listdir(subb_path)
                target_files = [os.path.join(subb_path, x) for x in dicom_list]

        if len(input_files) < 200:
            continue
        if len(target_files) % len(input_files) != 0:
            print(input_files[0])
            continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        for j in range(len(input_files)):
            input_path = input_files[j]  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
            dsA.SeriesDescription = 'CT'
            dsA.SeriesNumber = 2
            DCMA.append(dsA)
            loca[0, j] = dsA.SliceLocation  ###层面是否对应

        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(target_files)):
            target_path = target_files[j]
            dsB = pydicom.dcmread(target_path, force=True)  #
            dsB.SeriesDescription = 'CTA'
            # dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
            # dsB.InstanceNumber = dsA.InstanceNumber
            dsB.SeriesInstanceUID = dsbsi  # 区分
            dsB.SeriesNumber = 3  # 区分
            DCMB.append(dsB)
            locb[0, j] = dsB.SliceLocation

        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].SliceLocation
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].SliceLocation:
                sl = DCMB[b].SliceLocation
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号
        # outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
        # outb_dicom, new_spacing = resample(imgb, dsB)
        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        if outb_dicom.shape[2] % outa_dicom.shape[2] != 0:
            continue
        elif outb_dicom.shape[2] < 150:
            continue
        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)

        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        if "neck" in sub_path:
            be = outa_dicom.shape[2] - 35
        else:
            be = 0
            num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1

            # dsB.SeriesNumber = 3  # 区分
            dsB.SliceLocation -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.SliceLocation]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
            f1.writelines(file_patha + "\n")
        try:
            if float(dsA.AcquisitionTime) > float(dsB.AcquisitionTime):
                os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
                os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
                os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        except:
            if float(dsA.InstanceCreationTime) > float(dsB.InstanceCreationTime):
                os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
                os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
                os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
            print(sub_path)
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    excel_name = "/xy.xlsx"
    excel_path = os.path.dirname(os.path.abspath(__file__)) + excel_name
    df = pd.DataFrame(columns=["ID", "path", "date", "thickness/mm", "normal", "aneurysm", "stenosis"])
    df.to_excel(excel_path, index=False)
    df = pd.read_excel(excel_path, header=None)
    ds = pd.DataFrame(name)
    df = df.append(ds, ignore_index=True)
    df.to_excel(excel_path, index=False, header=False)
    f1.close()
    print('id numbers:', len(id))


def make_abd_process2b():
    path = "/media/yml/yml_data2/data_processed/FB/xy3/"  # neckk
    # path="../../../data/pv_data/jc/DICOM1/"
    output = "/media/yml/yml_data2/data_use/pv_data/xy3"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)
    ij = 1
    id = []
    name = []
    for sub_path in path_list:
        # if ij<23:
        #     ij=ij+1
        #     continue
        putpath = os.path.join(output, "PA" + str(ij))
        file_list = os.listdir(sub_path)
        # ncct=[]
        # cta=[]
        for file in file_list:
            if "CTA" not in file and "SE" not in file:
                subb_path = os.path.join(sub_path, file)
                # dicom_list=os.listdir(subb_path)
                # input_files=[os.path.join(subb_path, x) for x in dicom_list]
                new = subb_path.replace(file, 'SE0')
                os.rename(subb_path, new)

            elif "SE0" not in file:
                subb_path = os.path.join(sub_path, file)
                new = subb_path.replace(file, 'SE1')
                os.rename(subb_path, new)

        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('id numbers:', len(id))


# xm
def make_abd_process3():
    path = "/media/yml/yml_data2/data_raw/xm/xm2/"
    output = "/media/yml/yml_data2/data_processed/FB/xm/xm3/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2 and "PA" in root:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()

    f1 = open("xm.txt", "w")  # 564
    ij = 1
    id = []
    name = []
    for sub_path in path_list:
        # if ij<71:
        #     ij=ij+1
        #     continue
        putpath = os.path.join(output, sub_path.split("xm2/")[1])
        file_list = os.listdir(sub_path)
        # ncct=[]
        # cta=[]
        for file in file_list:
            fpath = os.path.join(sub_path, file)
            dicom_list = os.listdir(fpath)
            flag_path = os.path.join(fpath, dicom_list[0])
            dsA = pydicom.dcmread(flag_path, force=True)  # 读取头文件 InstanceNumber
            if "SRS" in file:
                if "A" in dsA.SeriesDescription:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    target_files = [os.path.join(subb_path, x) for x in dicom_list]

                else:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    input_files = [os.path.join(subb_path, x) for x in dicom_list]
            else:
                if "A" in file:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    target_files = [os.path.join(subb_path, x) for x in dicom_list]

                else:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    input_files = [os.path.join(subb_path, x) for x in dicom_list]

        if len(input_files) < 200:
            continue
        # if len(target_files)%len(input_files)!=0:
        #     print(input_files[0])
        #     continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        for j in range(len(input_files)):
            input_path = input_files[j]  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
            dsA.SeriesDescription = 'CT'
            dsA.SeriesNumber = 2
            DCMA.append(dsA)
            loca[0, j] = dsA.ImagePositionPatient[2]  ###层面是否对应

        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(target_files)):
            target_path = target_files[j]
            dsB = pydicom.dcmread(target_path, force=True)  #
            dsB.SeriesDescription = 'CTA'
            # dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
            dsB.InstanceNumber = dsA.InstanceNumber
            dsB.SeriesInstanceUID = dsbsi  # 区分
            dsB.SeriesNumber = 3  # 区分
            DCMB.append(dsB)
            locb[0, j] = dsB.ImagePositionPatient[2]

        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].ImagePositionPatient[2]
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].ImagePositionPatient[2]:
                sl = DCMA[a].ImagePositionPatient[2]
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].ImagePositionPatient[2]
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].ImagePositionPatient[2]:
                sl = DCMB[b].ImagePositionPatient[2]
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号
        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        # if outb_dicom.shape[2]%outa_dicom.shape[2]!=0:
        #     continue
        # elif outb_dicom.shape[2]<150:
        #     continue
        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)
        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        if "neck" in sub_path:
            be = outa_dicom.shape[2] - 35
        else:
            be = 0
            num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.ImagePositionPatient[2] -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.ImagePositionPatient[2]]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1
            file_patha = os.path.join(putpath1, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            f1.writelines(file_patha + "\n")

        for k in range(be, outb_dicom.shape[2]):
            # dsB.SeriesNumber = 3  # 区分
            dsB.ImagePositionPatient[2] -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.ImagePositionPatient[2]]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
        # try:
        #     if float(dsA.AcquisitionTime) > float(dsB.AcquisitionTime):
        #         os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
        #         os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
        #         os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        # except:
        #     if float(dsA.InstanceCreationTime) > float(dsB.InstanceCreationTime):
        #         os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
        #         os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
        #         os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        #     print(sub_path)
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('id numbers:', len(id))


def make_abd_process3b():
    path = "/media/yml/yml_data2/data_processed/FB/xm/xm2/"
    output = "/media/yml/yml_data2/data_use/pv_data/xm/xm3/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2 and "PA" in root:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()

    ij = 1
    id = []
    name = []
    for sub_path in path_list:
        # if ij<36:
        #     ij=ij+1
        #     continue
        putpath = os.path.join(output, sub_path.split("xm2/")[1])
        file_list = os.listdir(sub_path)
        # ncct=[]
        # cta=[]
        for file in file_list:
            fpath = os.path.join(sub_path, file)
            dicom_list = os.listdir(fpath)
            flag_path = os.path.join(fpath, dicom_list[0])
            dsA = pydicom.dcmread(flag_path, force=True)  # 读取头文件 InstanceNumber
            if "SRS" in file:
                if "A" in dsA.SeriesDescription:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    target_files = [os.path.join(subb_path, x) for x in dicom_list]

                else:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    input_files = [os.path.join(subb_path, x) for x in dicom_list]
            else:
                if "A" in file:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    target_files = [os.path.join(subb_path, x) for x in dicom_list]

                else:
                    subb_path = os.path.join(sub_path, file)
                    dicom_list = os.listdir(subb_path)
                    input_files = [os.path.join(subb_path, x) for x in dicom_list]

        if len(input_files) < 200:
            continue
        if len(target_files) % len(input_files) != 0:
            print(input_files[0])
            continue
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(input_files)))
        dsbsi = pydicom.uid.generate_uid()
        try:
            for j in range(len(input_files)):
                input_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
                target_path = os.path.join(sub_path, target_files[j])
                dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
                dsB = pydicom.dcmread(target_path,
                                      force=True)  # 读取头文件  1.2.840.113619.2.278.3.2831212038.186.1561204307.645
                if dsA.PatientID not in id:
                    id.append(dsA.PatientID)
                elif id[-1] != dsA.PatientID:  # 一个文件夹下应该为同一个人，如果出现不同的patient，则删除。
                    shutil.rmtree(sub_path)
                    print(sub_path)
                    continue
                dsA.SeriesDescription = 'CT'
                dsA.SeriesNumber = 2

                dsB.SeriesDescription = 'CTA'
                dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
                dsB.InstanceNumber = dsA.InstanceNumber
                dsB.SeriesInstanceUID = dsbsi  # 区分
                dsB.SeriesNumber = 3  # 区分
                DCMA.append(dsA)
                DCMB.append(dsB)
                loca[0, j] = dsA.ImagePositionPatient[2]  ###层面是否对应
                locb[0, j] = dsB.ImagePositionPatient[2]
        except:
            print(sub_path)
            continue

        idxa = np.argsort((loca))
        idxb = np.argsort((locb))
        # putpath1 = putpath + str(i) + '/SE0'
        # putpath2 = putpath + str(i) + '/SE1'
        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        try:
            for k in range(len(input_files)):  # 排序，保证命名和序列一致
                if not os.path.isdir(putpath1):
                    os.makedirs(putpath1)
                    os.makedirs(putpath2)
                file_patha = os.path.join(putpath1, str(k + 1) + '.dcm')
                file_pathb = os.path.join(putpath2, str(k + 1) + '.dcm')
                a = idxa[0, k]
                b = idxb[0, k]
                pydicom.dcmwrite(file_patha, DCMA[a])
                pydicom.dcmwrite(file_pathb, DCMB[b])
        except:
            print(sub_path)
            continue
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('id numbers:', len(id))


# xm 处理命名比较规范的数据
def make_abd_process4():
    path = "../../../data/pv_data/xm0/"
    output = "../../../data/pv_data/xm00/"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2 and "PA" in root:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()
    # random.shuffle(path_list)

    f1 = open("xm.txt", "w")  # 564
    ij = 1
    id = []
    name = []
    for sub_path in path_list:
        # if ij<25:
        #     ij=ij+1
        #     continue
        putpath = os.path.join(output, "PA" + str(ij))
        # putpath = os.path.join(output, sub_path.split("xm/")[1])
        file_list = os.listdir(sub_path)
        # ncct=[]
        # cta=[]
        for file in file_list:
            if "SE0" in file:
                subb_path = os.path.join(sub_path, file)
                dicom_list = os.listdir(subb_path)
                input_files = [os.path.join(subb_path, x) for x in dicom_list]
            else:
                subb_path = os.path.join(sub_path, file)
                dicom_list = os.listdir(subb_path)
                target_files = [os.path.join(subb_path, x) for x in dicom_list]

        if len(input_files) < 200:
            continue
        # if len(target_files)%len(input_files)!=0:
        #     print(input_files[0])
        #     continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        for j in range(len(input_files)):
            input_path = input_files[j]  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
            dsA.SeriesDescription = 'CT'
            dsA.SeriesNumber = 2
            dsA.WindowCenter = 40.0
            dsA.WindowWidth = 400.0
            DCMA.append(dsA)
            loca[0, j] = dsA.ImagePositionPatient[2]  # SliceLocation  ###层面是否对应

        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(target_files)):
            target_path = target_files[j]
            dsB = pydicom.dcmread(target_path, force=True)  #
            dsB.SeriesDescription = 'CTA'
            # dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
            # dsB.InstanceNumber = dsA.InstanceNumber
            dsB.SeriesInstanceUID = dsbsi  # 区分
            dsB.SeriesNumber = 3  # 区分
            dsB.WindowCenter = 40.0
            dsB.WindowWidth = 400.0
            DCMB.append(dsB)
            locb[0, j] = dsB.ImagePositionPatient[2]  # SliceLocation

        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].ImagePositionPatient[2]  # .SliceLocation
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].ImagePositionPatient[2]:
                sl = DCMA[a].ImagePositionPatient[2]
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].ImagePositionPatient[2]  # .SliceLocation
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].ImagePositionPatient[2]:
                sl = DCMB[b].ImagePositionPatient[2]
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号
        # imga[imga==0]=-2000 #可能会引入小圈圈
        # imgb[imgb==0]=-2000
        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        # if outb_dicom.shape[2]%outa_dicom.shape[2]!=0:
        #     continue
        if outb_dicom.shape[2] < 150:
            continue
        putpath1 = putpath + '/SE0'
        putpath2 = putpath + '/SE1'
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)

        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        if "neck" in sub_path:
            be = outa_dicom.shape[2] - 35
        else:
            be = 0
            num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.ImagePositionPatient[2] -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.ImagePositionPatient[2]]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1

            # dsB.SeriesNumber = 3  # 区分
            dsB.ImagePositionPatient[2] -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.ImagePositionPatient[2]]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
            f1.writelines(file_patha + "\n")
        # try:
        #     if float(dsA.AcquisitionTime) > float(dsB.AcquisitionTime):
        #         os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
        #         os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
        #         os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        # except:
        #     if float(dsA.InstanceCreationTime) > float(dsB.InstanceCreationTime):
        #         os.rename(sub_path, sub_path.replace("SE0", "SE33"))  # 0 3
        #         os.rename(sub_path.replace("SE0", "SE1"), sub_path)  # 1 0
        #         os.rename(sub_path.replace("SE0", "SE33"), sub_path.replace("SE0", "SE1"))  # 3 1
        #     print(sub_path)
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    # excel_name = "/xy.xlsx"
    # excel_path = os.path.dirname(os.path.abspath(__file__)) + excel_name
    # df = pd.DataFrame(columns=["ID","path", "date", "thickness/mm", "normal", "aneurysm", "stenosis"])
    # df.to_excel(excel_path, index=False)
    # df = pd.read_excel(excel_path, header=None)
    # ds = pd.DataFrame(name)
    # df = df.append(ds, ignore_index=True)
    # df.to_excel(excel_path, index=False, header=False)
    # f1.close()
    print('id numbers:', len(id))


# cq
def make_abd_process5():
    catch = "/media/wangtao/yml_data2/data_raw/p2/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)
    path = "/media/wangtao/yml_data2/data_raw/p2/cqq"
    output = "/media/wangtao/yml_data2/data_raw/p2/cqqq"  # 中间缓存
    # output2 = "E:\data_raw\p2\ccq"
    output_nii = "/media/wangtao/yml_data2/data_raw/p2/cq_nii"

    path_list = []
    num_i = 0
    for root, dirs, files in os.walk(path, topdown=False):
        list = os.listdir(root)
        if len(list) == 2:
            path = os.path.join(root, dirs[0])
            list2 = os.listdir(path)
            if len(list2) > 200:
                se0 = int(dirs[0].split("_")[-1])
                se1 = int(dirs[1].split("_")[-1])
                if se0 > se1:
                    patha = os.path.join(root, dirs[1])
                    pathb = os.path.join(root, dirs[0])
                    pathaa = patha.replace(dirs[1], "SE0")
                    pathbb = pathb.replace(dirs[0], "SE1")
                    os.rename(patha, pathaa)  # 0 3
                    os.rename(pathb, pathbb)  # 0 3
                elif se0 < se1:
                    patha = os.path.join(root, dirs[0])
                    pathb = os.path.join(root, dirs[1])
                    pathaa = patha.replace(dirs[0], "SE0")
                    pathbb = pathb.replace(dirs[1], "SE1")
                    os.rename(patha, pathaa)  # 0 3
                    os.rename(pathb, pathbb)  # 0 3

                if "nor" in patha:
                    putouta = os.path.join(output, "nor/PA" + str(num_i) + "/SE0")
                    putoutb = os.path.join(output, "nor/PA" + str(num_i) + "/SE1")
                elif "xz" in patha:
                    putouta = os.path.join(output, 'dis/xz/PA' + str(num_i) + "/SE0")
                    putoutb = os.path.join(output, "dis/xz/PA" + str(num_i) + "/SE1")
                elif "dml" in patha:
                    putouta = os.path.join(output, "dis/dml/PA" + str(num_i) + "/SE0")
                    putoutb = os.path.join(output, "dis/dml/PA" + str(num_i) + "/SE1")
                elif "jc" in patha:
                    putouta = os.path.join(output, "dis/jc/PA" + str(num_i) + "/SE0")
                    putoutb = os.path.join(output, "dis/jc/PA" + str(num_i) + "/SE1")
                elif "dmzyyh" in patha:
                    putouta = os.path.join(output, "dis/dmzyyh/PA" + str(num_i) + "/SE0")
                    putoutb = os.path.join(output, "dis/dmzyyh/PA" + str(num_i) + "/SE1")
                num_i = num_i + 1
                if os.path.exists(putouta):
                    pass
                else:
                    os.makedirs(putouta.split('/SE')[0])
                shutil.move(pathaa, putouta)
                shutil.move(pathbb, putoutb)
                path_list.append(putouta)
    path_list.sort()

    path_list = []
    for root, dirs, files in os.walk(output, topdown=False):
        list = os.listdir(root)
        if "SE0" in root:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()

    f1 = open("cq.txt", "w")  # 564
    ij = 1
    id = []
    name = []
    flag = 0
    for sub_path in path_list:
        # if ij<275:
        #     ij=ij+1
        #     continue
        input_files = os.listdir(sub_path)
        target_files = os.listdir(sub_path.replace("SE0", "SE1"))
        if len(input_files) < 200:
            continue
        if len(target_files) % len(input_files) != 0:
            print(sub_path)
            continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        for j in range(len(input_files)):
            input_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
            dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
            dsA.SeriesDescription = 'CT'
            dsA.SeriesNumber = 2
            DCMA.append(dsA)
            loca[0, j] = dsA.SliceLocation  ###层面是否对应

        dsbsi = pydicom.uid.generate_uid()
        for j in range(len(target_files)):
            target_path = os.path.join(sub_path.replace("SE0", "SE1"), target_files[j])
            dsB = pydicom.dcmread(target_path, force=True)  #
            dsB.SeriesDescription = 'CTA'
            # dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
            # dsB.InstanceNumber = dsA.InstanceNumber
            dsB.SeriesInstanceUID = dsbsi  # 区分
            dsB.SeriesNumber = 3  # 区分
            DCMB.append(dsB)
            locb[0, j] = dsB.SliceLocation

        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].SliceLocation
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].SliceLocation:
                sl = DCMB[b].SliceLocation
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号
        # outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
        # outb_dicom, new_spacing = resample(imgb, dsB)
        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        if outb_dicom.shape[2] % outa_dicom.shape[2] != 0:
            continue
        elif outb_dicom.shape[2] < 150:
            continue
        putpath1 = sub_path.replace("cqqq", "cq")
        putpath2 = sub_path.replace("cqqq", "cq").replace("SE0", "SE1")
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)

        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        be = 0
        num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1

            # dsB.SeriesNumber = 3  # 区分
            dsB.SliceLocation -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.SliceLocation]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
            f1.writelines(file_patha + "\n")
        se0output = os.path.join(catch, '0.nii.gz')
        dcm2nii_sitk(putpath1, se0output)
        se1output = os.path.join(catch, '1.nii.gz')
        dcm2nii_sitk(putpath2, se1output)
        registor(se0output, se1output)
        out_file_path = os.path.join(output_nii, putpath1.split("cq/")[1].split("SE")[0])  ####
        if not os.path.isdir(out_file_path):
            os.makedirs(out_file_path)

        read0 = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read0)
        out0 = sitk.GetImageFromArray(img_array0)
        sitk.WriteImage(out0, se0output)

        read1 = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read1)
        out1 = sitk.GetImageFromArray(img_array1)
        sitk.WriteImage(out1, se1output)
        x1, y1, z1 = img_array0.shape
        x2, y2, z2 = img_array1.shape
        if x1 != x2 or y1 != y2 or z1 != z2:
            print(sub_path)
        shutil.move(se0output, out_file_path + "/0.nii.gz")
        shutil.move(se1output, out_file_path + "/1.nii.gz")
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('id numbers:', len(id))


def make_abd_process5b():
    catch = "/media/wangtao/yml_data2/data_raw/p2/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)
    path = "/media/wangtao/yml_data2/data_raw/p2/lz"
    output_nii = "/media/wangtao/yml_data2/data_use/p2_nii/external/lz"

    f = open("registor_list.txt")  # test train！！  list  filter
    path_list = []
    for line in f.readlines():
        if "SE0" in line:
            path_list.append(line.split('\n')[0])
    path_list.sort()
    ij = 0
    for sub_path in path_list:
        # if ij<6:
        #     ij=ij+1
        #     continue
        sub_path = "/media/wangtao/yml_data2/data_raw/p2/lzz/dis/dml/PA0/SE0"
        putpath1 = sub_path  # sub_path.replace("cqqq","cq")
        putpath2 = sub_path.replace("SE0", "SE1")
        # if not os.path.isdir(putpath1):
        #     os.makedirs(putpath1)
        #     os.makedirs(putpath2)

        # save to dicom file
        try:
            se0output = os.path.join(catch, '0.nii.gz')
            dcm2nii_sitk(putpath1, se0output)
            se1output = os.path.join(catch, '1.nii.gz')
            dcm2nii_sitk(putpath2, se1output)
            registor(se0output, se1output)
        except:
            print(sub_path)
            continue
        out_file_path = os.path.join(output_nii, putpath1.split("lzz/")[1].split("SE")[0])  ####
        if not os.path.isdir(out_file_path):
            os.makedirs(out_file_path)

        read0 = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read0)
        out0 = sitk.GetImageFromArray(img_array0)
        sitk.WriteImage(out0, se0output)

        read1 = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read1)
        out1 = sitk.GetImageFromArray(img_array1)
        sitk.WriteImage(out1, se1output)
        shutil.move(se0output, out_file_path + "/0.nii.gz")
        shutil.move(se1output, out_file_path + "/1.nii.gz")
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('finished:!')

# lz
def unmatchdata_process(sub_path, output_nii, catch):
    putpath1 = sub_path  # sub_path.replace("cqqq","cq")
    putpath2 = sub_path.replace("SE0", "SE1")
    se0output = os.path.join(catch, '0.nii.gz')
    dcm2nii_sitk(putpath1, se0output)
    se1output = os.path.join(catch, '1.nii.gz')
    dcm2nii_sitk(putpath2, se1output)
    registor(se0output, se1output)
    read0 = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
    img_array0 = sitk.GetArrayFromImage(read0)
    out0 = sitk.GetImageFromArray(img_array0)
    sitk.WriteImage(out0, se0output)

    read1 = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
    img_array1 = sitk.GetArrayFromImage(read1)
    out1 = sitk.GetImageFromArray(img_array1)
    sitk.WriteImage(out1, se1output)
    x1, y1, z1 = img_array0.shape
    x2, y2, z2 = img_array1.shape
    if x1 != x2 or y1 != y2 or z1 != z2:
        print("Slice dose not equal")
    out_file_path = os.path.join(output_nii, putpath1.split("lzz/")[1].split("SE")[0])  ####
    if not os.path.isdir(out_file_path):
        os.makedirs(out_file_path)
    shutil.move(se0output, out_file_path + "/0.nii.gz")
    shutil.move(se1output, out_file_path + "/1.nii.gz")
    # print(sub_path)


def make_abd_process6():
    catch = "/media/wangtao/yml_data2/data_raw/p2/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)
    path = "/media/wangtao/yml_data2/data_raw/p2/lz"
    output = "/media/wangtao/yml_data2/data_raw/p2/lzz"  # 中间缓存
    # output_nii = "/media/wangtao/yml_data2/data_raw/p2/nii/lz"
    output_nii = "/media/wangtao/yml_data2/data_use/p2_nii/external/lz"

    # path_list = []
    # num_i = 42
    # for root, dirs, files in os.walk(path, topdown=False):
    #     file_list = os.listdir(root)
    #     if len(file_list) == 2:
    #         path = os.path.join(root, dirs[0])
    #         list2 = os.listdir(path)
    #         if len(list2) > 200:
    #             if "平扫" in dirs[1]:
    #                 patha = os.path.join(root, dirs[1])
    #                 pathb = os.path.join(root, dirs[0])
    #                 pathaa = patha.replace(dirs[1], "SE0")
    #                 pathbb = pathb.replace(dirs[0], "SE1")
    #                 os.rename(patha, pathaa)  # 0 3
    #                 os.rename(pathb, pathbb)  # 0 3
    #             elif "平扫" in dirs[0]:
    #                 patha = os.path.join(root, dirs[0])
    #                 pathb = os.path.join(root, dirs[1])
    #                 pathaa = patha.replace(dirs[0], "SE0")
    #                 pathbb = pathb.replace(dirs[1], "SE1")
    #                 os.rename(patha, pathaa)  # 0 3
    #                 os.rename(pathb, pathbb)  # 0 3
    #
    #             if "nor" in patha:
    #                 putouta = os.path.join(output, "nor/PA" + str(num_i) + "/SE0")
    #                 putoutb = os.path.join(output, "nor/PA" + str(num_i) + "/SE1")
    #             elif "xz" in patha:
    #                 putouta = os.path.join(output, 'dis/xz/PA' + str(num_i) + "/SE0")
    #                 putoutb = os.path.join(output, "dis/xz/PA" + str(num_i) + "/SE1")
    #             elif "dml" in patha:
    #                 putouta = os.path.join(output, "dis/dml/PA" + str(num_i) + "/SE0")
    #                 putoutb = os.path.join(output, "dis/dml/PA" + str(num_i) + "/SE1")
    #             elif "jc" in patha:
    #                 putouta = os.path.join(output, "dis/jc/PA" + str(num_i) + "/SE0")
    #                 putoutb = os.path.join(output, "dis/jc/PA" + str(num_i) + "/SE1")
    #             elif "dmzyyh" in patha:
    #                 putouta = os.path.join(output, "dis/dmzyyh/PA" + str(num_i) + "/SE0")
    #                 putoutb = os.path.join(output, "dis/dmzyyh/PA" + str(num_i) + "/SE1")
    #             num_i = num_i + 1
    #             if os.path.exists(putouta):
    #                 pass
    #             else:
    #                 os.makedirs(putouta.split('/SE')[0])
    #             shutil.move(pathaa, putouta)
    #             shutil.move(pathbb, putoutb)
    #             path_list.append(putouta)
    # path_list.sort()

    path_list = []
    for root, dirs, files in os.walk(output, topdown=False):
        list = os.listdir(root)
        if "SE0" in root:
            # if"Abdomen" not in root:
            path_list.append(root)
    path_list.sort()

    f1 = open("lz.txt", "w")  # 564
    ij = 1
    id = []
    name = []
    flag = 0
    for sub_path in path_list:
        if ij < 140:
            ij = ij + 1
            continue
        input_files = os.listdir(sub_path)
        target_files = os.listdir(sub_path.replace("SE0", "SE1"))
        if len(input_files) < 200:
            continue
        if len(target_files) % len(input_files) != 0:
            try:
                unmatchdata_process(sub_path, output_nii, catch)
                ij = ij + 1
                if ij % 10 == 0:
                    print('numbers:', ij)
                continue
            except:
                print("registor wrong")
                print(sub_path)
                continue
        # get file list
        DCMA = []  # np.zeros((512,512,len(data_files))) #（4）用于解决命名和序列是否一致
        DCMB = []  # np.zeros((512,512,len(data_files)))
        loca = np.zeros((1, len(input_files)))
        locb = np.zeros((1, len(target_files)))
        try:
            for j in range(len(input_files)):
                input_path = os.path.join(sub_path, input_files[j])  # 直接将文件中按行读到list里，效果与方法2一样
                dsA = pydicom.dcmread(input_path, force=True)  # 读取头文件 InstanceNumber
                dsA.SeriesDescription = 'CT'
                dsA.SeriesNumber = 2
                DCMA.append(dsA)
                loca[0, j] = dsA.SliceLocation  ###层面是否对应

            dsbsi = pydicom.uid.generate_uid()
            for j in range(len(target_files)):
                target_path = os.path.join(sub_path.replace("SE0", "SE1"), target_files[j])
                dsB = pydicom.dcmread(target_path, force=True)  #
                dsB.SeriesDescription = 'CTA'
                # dsB.AcquisitionNumber = dsA.AcquisitionNumber  # 保持一致
                # dsB.InstanceNumber = dsA.InstanceNumber
                dsB.SeriesInstanceUID = dsbsi  # 区分
                dsB.SeriesNumber = 3  # 区分
                DCMB.append(dsB)
                locb[0, j] = dsB.SliceLocation
        except:
            continue
        # read data
        file_n1 = len(np.unique(loca))
        file_n2 = len(np.unique(locb))
        imga = np.zeros([file_n1, 512, 512])
        imgb = np.zeros([file_n2, 512, 512])
        idxa = np.argsort((-loca))
        idxb = np.argsort((-locb))
        # idxa = np.argsort((loca))
        # idxb = np.argsort((locb))

        kk = 1
        sl = DCMA[idxa[0, 0]].SliceLocation
        for k in range(file_n1):  # 排序，保证命名和序列一致
            a = idxa[0, k]
            if sl == DCMA[a].SliceLocation:
                sl = DCMA[a].SliceLocation
                kk = kk - 1
            try:
                imga[kk, :, :] = DCMA[a].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        kk = 1
        sl = DCMB[idxb[0, 0]].SliceLocation
        for k in range(file_n2):  # 排序，保证命名和序列一致
            b = idxb[0, k]
            if sl == DCMB[b].SliceLocation:
                sl = DCMB[b].SliceLocation
                kk = kk - 1
            try:
                imgb[kk, :, :] = DCMB[b].pixel_array.astype(np.int16)
            except:
                continue
            kk = kk + 1

        # resample
        dsA = DCMA[idxa[0, 0]]
        dsB = DCMB[idxb[0, 0]]
        dsA.BitsStored = 16
        dsB.BitsStored = 16
        dsA.PixelRepresentation = 1  # 有符号，将无符号转换为有符号
        dsB.PixelRepresentation = 1  # 有符号
        # outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
        # outb_dicom, new_spacing = resample(imgb, dsB)
        try:
            outa_dicom, new_spacing = resample(imga, dsA)  ###不需要
            outb_dicom, new_spacing = resample(imgb, dsB)
        except:
            continue
        if outb_dicom.shape[2] % outa_dicom.shape[2] != 0:
            continue
        elif outb_dicom.shape[2] < 150:
            continue
        putpath1 = sub_path.replace("lzz", "lz")
        putpath2 = sub_path.replace("lzz", "lz").replace("SE0", "SE1")
        if not os.path.isdir(putpath1):
            os.makedirs(putpath1)
            os.makedirs(putpath2)

        # save to dicom file
        dsbsi = pydicom.uid.generate_uid()
        be = 0
        num = outa_dicom.shape[2]
        for k in range(be, outa_dicom.shape[2]):
            dsA.SliceLocation -= new_spacing[0]
            dsA.ImagePositionPatient = [dsA.ImagePositionPatient[0], dsA.ImagePositionPatient[1], dsA.SliceLocation]
            newimga = outa_dicom[:, :, k].astype(np.int16)
            dsA.PixelData = newimga.tobytes()
            dsA.PixelSpacing = new_spacing[1:2]
            dsA.SliceThickness = new_spacing[0]
            dsA.SeriesNumber = 2
            dsA.InstanceNumber = k + 1

            # dsB.SeriesNumber = 3  # 区分
            dsB.SliceLocation -= new_spacing[0]
            dsB.ImagePositionPatient = [dsB.ImagePositionPatient[0], dsB.ImagePositionPatient[1], dsB.SliceLocation]
            dsB.SeriesInstanceUID = dsbsi  # 区分
            newimgb = outb_dicom[:, :, k].astype(np.int16)
            dsB.PixelData = newimgb.tobytes()
            dsB.PixelSpacing = new_spacing[1:2]
            dsB.SliceThickness = new_spacing[0]
            dsB.SeriesNumber = 3  # 区分
            dsB.InstanceNumber = k + 1

            file_patha = os.path.join(putpath1, 'IM' + str(k))
            file_pathb = os.path.join(putpath2, 'IM' + str(k))
            pydicom.dcmwrite(file_patha, dsA)
            pydicom.dcmwrite(file_pathb, dsB)
            # 随机划分子数据集
            f1.writelines(file_patha + "\n")
        se0output = os.path.join(catch, '0.nii.gz')
        dcm2nii_sitk(putpath1, se0output)
        se1output = os.path.join(catch, '1.nii.gz')
        dcm2nii_sitk(putpath2, se1output)
        try:
            registor(se0output, se1output)
        except:
            print(sub_path)
            continue
        out_file_path = os.path.join(output_nii, putpath1.split("lz/")[1].split("SE")[0])  ####
        if not os.path.isdir(out_file_path):
            os.makedirs(out_file_path)

        read0 = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read0)
        out0 = sitk.GetImageFromArray(img_array0)
        sitk.WriteImage(out0, se0output)

        read1 = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read1)
        out1 = sitk.GetImageFromArray(img_array1)
        sitk.WriteImage(out1, se1output)
        x1, y1, z1 = img_array0.shape
        x2, y2, z2 = img_array1.shape
        if x1 != x2 or y1 != y2 or z1 != z2:
            print(sub_path)
        shutil.move(se0output, out_file_path + "/0.nii.gz")
        shutil.move(se1output, out_file_path + "/1.nii.gz")
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('id numbers:', len(id))

def registor(patha, pathb):
    # 本例子展示的是将CTA配准至CT
    fixed_image = itk.imread(patha, itk.F)  # 这里itk.F是itk的一种图像类型，需要加itk.F代码才能在该类型精度下进行配准
    moving_image = itk.imread(pathb, itk.F)

    # 配准参数：rigid+affine+bspline
    # 注：只使用rigid也可以，加上affine+bspline允许一些弹性变形变化，会配得更好
    # 但在某些情况下使用affine+bspline会使得moving图像变形严重
    # parameter_object = itk.ParameterObject.New()
    parameter_object = itk.ParameterObject()
    parameter_map_rigid = parameter_object.GetDefaultParameterMap('rigid')
    parameter_object.AddParameterMap(parameter_map_rigid)
    abc = parameter_object
    try:
        parameter_map_affine = abc.GetDefaultParameterMap('affine')
        abc.AddParameterMap(parameter_map_affine)
        parameter_object = abc

        parameter_map_bspline = abc.GetDefaultParameterMap('bspline')
        abc.AddParameterMap(parameter_map_bspline)
        parameter_object = abc
    except:
        print("registor have warning!")

    # result_registered_image就是配准之后的ct图像，可以直接输出保存
    # 同时配准结束后还会输出result_transform_parameters，是配准的变换参数，可以用它来变换标签
    result_registered_image, result_transform_parameters = itk.elastix_registration_method(
        fixed_image, moving_image, parameter_object=parameter_object)
    itk.imwrite(result_registered_image, pathb)


# dicom to nii
def dcm2nii_sitk(path_read, path_save):
    reader = sitk.ImageSeriesReader()
    seriesIDs = reader.GetGDCMSeriesIDs(path_read)
    N = len(seriesIDs)
    lens = np.zeros([N])
    for i in range(N):
        dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[i])
        lens[i] = len(dicom_names)
    N_MAX = np.argmax(lens)
    dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[N_MAX])
    reader.SetFileNames(dicom_names)
    image = reader.Execute()
    sitk.WriteImage(image, path_save)
    # shutil.rmtree(path_read)


# 依据粗略标注的血液分割mask，标注钙化斑块，精细化标注
def remove_some_region(img, min_point=3, max_point=300):  # 80
    img_label, num = measure.label(img, connectivity=2, return_num=True)  # img为二值图，输出二值图像中所有的连通域
    props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等
    resMatrix = np.zeros(img_label.shape)
    for i in range(0, len(props)):
        are = props[i].area
        if are > min_point:
            if are < max_point:
                # dia=int(np.sqrt(are)/2)
                tmp = (img_label == i + 1).astype(np.uint8)
                # kernel = np.ones((dia, dia), np.uint8)
                # tmp = cv2.dilate(tmp, kernel)#[-1 1]
                resMatrix += tmp  # 组合所有符合条件的连通域
    resMatrix *= 1
    resMatrix[resMatrix > 1] = 1  #
    return resMatrix


# 提取钙化mask使用。血流、钙化斑块
def vessel_to_Calcified():
    # path = "../../../data/p2_nii/p1"
    # path = "../../../data/p1_nii"#~/yml/data/pv_nii/xm/xm1/dis/dml/PA1
    path = "../../../data/p2_s/xy/xy1/dis/dml/PA35/SE0"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path:
                path_list.append(path)
    i = 1
    for se1output in path_list:
        # if i<181:
        #     i = i + 1
        #     # break
        #     continue
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)
        # img_array_w2 = to_windowdata(img_array1, 80, 10)
        # img_array_w3 = to_windowdata(img_array1, 140, 10)
        img_array_w2 = to_windowdata(img_array1, 65, 10)
        img_array_w3 = to_windowdata(img_array1, 130, 10)

        se2output = se1output.replace("0.nii.gz", "2.nii.gz")  # 血液分割标签
        img2 = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array2 = sitk.GetArrayFromImage(img2)
        # print(img_array2.max())

        kernel1 = np.ones((3, 3), np.uint8)  # 用于中间过程
        # kernel2 = np.ones((40, 40), np.uint8)#用于血液mask
        if "dml" in se1output:
            kernel2 = np.ones((60, 60), np.uint8)
        else:
            kernel2 = np.ones((40, 40), np.uint8)
        new_array = np.full_like(img_array2, 1)
        for k in range(img_array2.shape[0]):
            img_cta2 = img_array_w2[k, :, :]  # ct
            img_cta22 = remove_some_region(img_cta2)  #
            img_cta3 = img_array_w3[k, :, :]
            img_cta33 = remove_some_region(img_cta3)
            img_cta44 = img_cta22 + img_cta33
            img_cta44 = np.where(img_cta44 > 1, 1, 0)
            img_cta55 = cv2.dilate(np.int16(img_cta44), kernel1)
            img_label2 = cv2.dilate(img_array2[k, :, :], kernel2)
            img_cta66 = img_cta55 * img_label2
            img_cta77 = cv2.dilate(np.int16(img_cta66), kernel1)
            img_cal = img_cta77 * img_cta22  # 钙化斑块
            img_cal = np.where(img_cal > 0.5, 1, 0)
            img_blood = img_array2[k, :, :]
            tot = img_cal + img_blood
            dif = np.where(tot > 1, 1, 0)
            img_blood = img_blood - dif
            # new_array[k,:,:] = img_cal+ img_blood*2 # [-1 1]
            new_array[k, :, :] = img_cal * 2 + img_blood

        out = sitk.GetImageFromArray(new_array)
        os.remove(se2output)
        sitk.WriteImage(out, se2output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


# 调试使用
def vessel_to_Calcified_test():
    path = "../../../data/p1_nii/dis/dml/fb/DICOM0/PA4"
    # path="../../../data/pv_nii/xy/xy1/dis/jc/PA1"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path:
                path_list.append(path)
    i = 1
    for se1output in path_list:
        sl = 219  # 224-223
        # read = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        # img_array0 = sitk.GetArrayFromImage(read)
        #
        # se1output = se0output.replace("0.nii.gz", "1.nii.gz")
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)
        img_array_w1 = to_windowdata(img_array1, 40, 300)
        img_array_w2 = to_windowdata(img_array1, 65, 10)
        img_array_w3 = to_windowdata(img_array1, 130, 10)

        img_cta = img_array_w1[sl, :, :]
        img_cta2 = img_array_w2[sl, :, :]
        img_cta22 = remove_some_region(img_cta2)  #
        img_cta3 = img_array_w3[sl, :, :]
        img_cta33 = remove_some_region(img_cta3)
        img_cta44 = img_cta22 + img_cta33
        img_cta44 = np.where(img_cta44 > 1, 1, 0)
        kernel1 = np.ones((3, 3), np.uint8)
        img_cta55 = cv2.dilate(np.int16(img_cta44), kernel1)

        se2output = se1output.replace("0.nii.gz", "2.nii.gz")
        img2 = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array2 = sitk.GetArrayFromImage(img2)
        img_label1 = img_array2[sl, :, :]

        # kernel2 = np.ones((60, 60), np.uint8)
        if "dml" in se1output:
            kernel2 = np.ones((60, 60), np.uint8)
        else:
            kernel2 = np.ones((40, 40), np.uint8)
        new_array = np.full_like(img_array2, 1)
        for k in range(img_array2.shape[0]):
            new_array[k, :, :] = cv2.dilate(img_array2[k, :, :], kernel2)  # [-1 1]
        img_label2 = new_array[sl, :, :]
        img_cta66 = img_cta55 * img_label2
        # img_cta66 = np.where(img_cta66 > 0.5, 1, 0)
        kernel3 = np.ones((3, 3), np.uint8)
        img_cta77 = cv2.dilate(np.int16(img_cta66), kernel3)
        img_cta88 = img_cta77 * img_cta22
        # image2 = image2 * img_array2
        # image2 = remove_small_points(image2, 30)

        plt.subplot(3, 3, 1)
        plt.imshow(img_cta, cmap='gray')  # ,vmin=0,vmax=255
        plt.imshow(img_label1, alpha=0.5, cmap='Greens')

        plt.subplot(3, 3, 2)
        plt.imshow(img_cta22, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 3)
        plt.imshow(img_cta3, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 4)
        plt.imshow(img_cta44, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 5)
        plt.imshow(img_cta55, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 6)
        plt.imshow(img_cta66, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 7)
        plt.imshow(img_cta77, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 8)
        plt.imshow(img_cta88, cmap='gray')  # ,vmin=0,vmax=255

        plt.subplot(3, 3, 9)
        # img_cta[ :, :][img_label2[:, :] > 0] = 0.1
        # plt.imshow(img_cta)
        plt.imshow(img_cta, cmap='gray')
        plt.imshow(img_cta88, alpha=0.8, cmap='Greens')  # img_label2
        plt.show()
        # img_label, num = measure.label(img1, connectivity=2, return_num=True)  # 输出二值图像中所有的连通域
        # props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等

        out = sitk.GetImageFromArray(img_array2)
        se3output = se1output.replace("1.nii.gz", "33.nii.gz")
        # os.remove(se2output)
        sitk.WriteImage(out, se3output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


# 剔除一些因为配准变形的数据
# 配准后的数据存在空，使得分割标注后的数据存在空白，这里需要注意下。目前还为发现配准后CTA的分割层面变少的情况
def remove_s_slices():
    path = "../../../data/p2_s/xy/xy1/dis/dml/PA35/SE0"
    # path="../../../data/pv_nii"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            # if "0.nii.gz" in path and "aug" in path:
            if "0.nii.gz" in path:
                path_list.append(path)
    ii = 1
    for se0output in path_list:
        se1output = se0output.replace("0.nii.gz", "1.nii.gz")
        img = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array = sitk.GetArrayFromImage(img)
        le = img_array.shape[0] - 1
        le1 = 2
        le2 = le
        # aa=img_array[0:1,:,:]
        for k in range(0, 12):
            img1 = img_array[k, :, :]
            img1 = np.where(img1 == 0, 1, 0)
            img_label, num = measure.label(img1, connectivity=2, return_num=True)  # 输出二值图像中所有的连通域
            props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等
            for i in range(0, len(props)):
                are = props[i].area
                if are > 500:
                    le1 = k + 1
                    continue
        for kk in range(le - 12, le):
            if le2 < le - 1:
                break
            img2 = img_array[kk, :, :]
            img2 = np.where(img2 == 0, 1, 0)
            img_label, num = measure.label(img2, connectivity=2, return_num=True)  # 输出二值图像中所有的连通域
            props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等
            for j in range(0, len(props)):
                are = props[j].area
                if are > 500:
                    le2 = kk + 1
                    break
            # le = img_array.shape[0] - 3
        # le1=80
        img_array = img_array[le1:le2, :, :]
        out = sitk.GetImageFromArray(img_array)
        os.remove(se1output)
        sitk.WriteImage(out, se1output)

        # os.remove(se0output)
        img = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array = sitk.GetArrayFromImage(img)
        img_array = img_array[le1:le2, :, :]
        out = sitk.GetImageFromArray(img_array)
        os.remove(se0output)
        sitk.WriteImage(out, se0output)

        try:
            read = se0output.replace("0.nii.gz", "2.nii.gz")
            img = sitk.ReadImage(read, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            img_array = sitk.GetArrayFromImage(img)
            img_array = img_array[le1:le2, :, :]
            out = sitk.GetImageFromArray(img_array)
            os.remove(read)
            sitk.WriteImage(out, read)
            # os.remove(read)
        except:
            continue

        try:
            read = se0output.replace("0.nii.gz", "3.nii.gz")
            img = sitk.ReadImage(read, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            img_array = sitk.GetArrayFromImage(img)
            img_array = img_array[le1:le2, :, :]
            out = sitk.GetImageFromArray(img_array)
            os.remove(read)
            sitk.WriteImage(out, read)
        except:
            continue

        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)


def caculate_weight_region(img):  # 80 input is mask label
    img = np.where(img > 0, 1, 0)
    all_w = 512 * 512 - img.sum()
    img_label, num = measure.label(img, connectivity=2, return_num=True)  # img为二值图，输出二值图像中所有的连通域
    props = measure.regionprops(img_label)  # 输出连通域的属性，包括面积等
    resMatrix = np.zeros(img_label.shape)
    mean_w = all_w * 3 / len(props)  # 每个连通区域总权值相同
    for i in range(0, len(props)):
        are = props[i].area
        w = mean_w / are
        # w=(0.999)**(are)#衰减函数使得mask 连通区域面积越大，权值越小。
        tmp = (img_label == i + 1).astype(np.uint8)
        resMatrix += tmp * w  # 组合所有符合条件的连通域
    # resMatrix *= 5000
    return resMatrix


def get_sample_map():
    # 有何作用？
    # path = "../../../data/p1_nii/dis/xz/fb/DICOM0/PA5"
    path = "../../../data/p1_m"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "1.nii.gz" in path:
                path_list.append(path)
    i = 1
    for se1output in path_list:
        try:
            # if i<181:
            #     i = i + 1
            #     # break
            #     continue
            read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            img_array1 = sitk.GetArrayFromImage(read)
            img_array_w = to_windowdata(img_array1, 75, 400)
            img_array_w = np.where(img_array_w > 0, 1, 0)

            se2output = se1output.replace("1.nii.gz", "2.nii.gz")  # 血液分割标签
            img2 = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            img_array2 = sitk.GetArrayFromImage(img2)

            new_array = np.full_like(img_array2, 1)
            for k in range(img_array2.shape[0]):
                mask = img_array2[k, :, :]  # ct
                mask2 = caculate_weight_region(mask)  #
                new_array[k, :, :] = mask2  # [-1 1]
            new_array = new_array + img_array_w
            out = sitk.GetImageFromArray(new_array)
            # os.remove(se2output)
            se3output = se1output.replace("1.nii.gz", "4.nii.gz")
            sitk.WriteImage(out, se3output)
        except:
            print(se1output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


def check_slice_num():
    path = "../../../data/p2_nii"
    # path="../../../data/pv_nii/xy/xy1/dis/jc/PA1"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path:
                path_list.append(path)
    i = 1
    for se1output in path_list:
        # read = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        # img_array0 = sitk.GetArrayFromImage(read)
        #
        # se1output = se0output.replace("0.nii.gz", "1.nii.gz")
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)
        le = img_array1.shape[0]
        if le < 150:
            print(le)
            print(se1output)
        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


# 划分训练-验证-测试数据
def split():
    # path="../../../data/p1_nii/dis/xz/"#CT_CTA disease
    # path = "../../../data/p2_nii/"  # CT_CTA disease p1_m
    path = "../../../data/p2_s/"  #
    path_list = []
    aug_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        if "0.nii.gz" in files:
            path_list.append(root)

        # if "PA" in root and "hnnk" not in root:#将hnnk作为外部验证集
        #     if "0.nii.gz" not in files:
        #         continue
        #     if "aug" not in root:
        #         if "nor" in root or "dmzyyh" in root:
        #             path_list.append(root)
        #     else:
        #         aug_list.append(root)

    random.seed(0)
    random.shuffle(path_list)
    cross = 5
    # ff = open("p1_data.txt", "w")#564
    a = int(len(path_list) / cross)
    for i in range(cross):
        if i == 1:
            break
        train = "trainn" + str(i) + ".txt"
        # val = "abd_val" + str(i) + ".txt"
        test = "testn" + str(i) + ".txt"
        f1 = open(train, "w")  # 564
        f2 = open(test, "w")  # 188
        for j in range(len(path_list)):
            input_files = os.listdir(path_list[j])  # SE0 file list
            file_path = os.path.join(path_list[j], "0.nii.gz")
            # ff.writelines(file_path)
            if j >= (cross - 1 - i) * a and j < (cross - i) * a:
                f2.writelines(file_path + "\n")
            else:
                f1.writelines(file_path + "\n")

        for j in range(len(aug_list)):
            file_path = os.path.join(aug_list[j], "0.nii.gz")
            f1.writelines(file_path + "\n")
    # ff.close()
    f1.close()  # 关
    f2.close()


############# augument #############
def centerCrop(image, output_size):
    if image.shape[0] <= output_size[0] or image.shape[1] <= output_size[1] or image.shape[2] <= output_size[2]:
        pw = max((output_size[0] - image.shape[0]) // 2 + 3, 0)
        ph = max((output_size[1] - image.shape[1]) // 2 + 3, 0)
        pd = max((output_size[2] - image.shape[2]) // 2 + 3, 0)
        image = np.pad(image, [(pw, pw), (ph, ph), (pd, pd)], mode='edge')
        # label = np.pad(label, [(pw, pw), (ph, ph), (pd, pd)], mode='constant', constant_values=0)

    (w, h, d) = image.shape
    w1 = int(round((w - output_size[0]) / 2.))
    h1 = int(round((h - output_size[1]) / 2.))
    d1 = int(round((d - output_size[2]) / 2.))
    # print(image.shape, output_size, get_center(label), w1, h1, d1)
    image = image[w1:w1 + output_size[0], h1:h1 + output_size[1], d1:d1 + output_size[2]]
    # label = label[w1:w1 + output_size[0], h1:h1 + output_size[1], d1:d1 + output_size[2]]
    return image


def padding_and_cropping(image, target_shape, big_pad=True,
                         borderType=cv2.BORDER_REPLICATE):
    # 目标尺寸大小
    ph, pw = target_shape
    # ph+=19*2
    # pw+=19*2
    # 原始图片尺寸
    h, w, _ = image.shape
    if big_pad and ph > h and pw > w:  # 以原图为中心进行边缘填充
        top = bottom = (ph - h) // 2  # 获取上、下填充尺寸
        top += (ph - h) % 2  # 为保证目标大小，无法整除则上+1
        left = right = (pw - w) // 2
        left += (pw - w) % 2  # 为保证目标大小，同理左上+1
        # image_padded = cv2.copyMakeBorder(image, top, bottom, left, right,
        #                                   borderType=borderType)
        # image_padded = cv2.copyMakeBorder(image, top, bottom, left, right, borderType, value = 0)

        pad_x_tuple = (top, top)
        pad_y_tuple = (left, left)
        pad_z_tuple = (0, 0)
        image_padded = np.pad(image, (pad_x_tuple, pad_y_tuple, pad_z_tuple),
                              mode='constant', constant_values=-2000)

    else:  # 最小比例缩放填充（大尺寸：高/宽比例变化较大的将被填充，小尺寸反之）
        # 计算缩放后图片尺寸
        image_padded = centerCrop(image, [512, 512, image.shape[2]])

    return image_padded

def resize_image_with_crop_or_pad(image, img_size=(512, 512), **kwargs):
    """Image resizing. Resizes image by cropping or padding dimension
     to fit specified size.
    Args:
        image (np.ndarray): image to be resized
        img_size (list or tuple): new image size
        kwargs (): additional arguments to be passed to np.pad
    Returns:
        np.ndarray: resized image
    """
    h, w, d = image.shape
    img_size = (img_size[0], img_size[1], d)
    assert isinstance(image, (np.ndarray, np.generic))
    assert (image.ndim - 1 == len(img_size) or image.ndim == len(img_size)), 'Example size doesnt fit image size'

    # Get the image dimensionality
    rank = len(img_size)

    # Create placeholders for the new shape
    from_indices = [[0, image.shape[dim]] for dim in range(rank)]
    to_padding = [[0, 0] for dim in range(rank)]

    slicer = [slice(None)] * rank

    # For each dimensions find whether it is supposed to be cropped or padded
    for i in range(rank):
        if image.shape[i] < img_size[i]:
            to_padding[i][0] = (img_size[i] - image.shape[i]) // 2
            to_padding[i][1] = img_size[i] - image.shape[i] - to_padding[i][0]
        else:
            from_indices[i][0] = int(np.floor((image.shape[i] - img_size[i]) / 2.))
            from_indices[i][1] = from_indices[i][0] + img_size[i]

        # Create slicer object to crop or leave each dimension
        slicer[i] = slice(from_indices[i][0], from_indices[i][1])

    # Pad the cropped image to extend the missing dimension
    # return np.pad(image[slicer], to_padding, **kwargs)
    return np.pad(image[tuple(slicer)], to_padding, **kwargs)

def cq_statistic():
    xls_file = './cq_statistic/total.xlsx'
    df = pd.read_excel(xls_file, engine='openpyxl')

    # 统计男女数量
    gender_counts = df['性别'].value_counts()
    print("男女数量统计:")
    print(gender_counts)

    # 转换年龄为数值型，忽略含有非数字的行
    # df['年龄'] = pd.to_numeric(df['年龄'], errors='coerce')
    df['年龄'] = df['年龄'].str.replace('岁', '').astype(int)
    # 计算年龄的四分位数
    age_quartiles = df['年龄'].dropna().quantile([.25, .5, .75])
    print("\n年龄四分位数统计:")
    print(age_quartiles)
    a=1

def lz_statistic():
    path = "/media/bit301/data/yml/data/original/p2/lz"
    # path = "../../../data/p1_nii"#~/yml/data/pv_nii/xm/xm1/dis/dml/PA1
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "CTA" in path:
                # if "xz" in path and "aug" not in path:
                path_list.append(path)
    f1 = open("statistic_lz.txt", "w")
    path_list.sort()
    path_list= natsorted(path_list)
    gender={'male':0,'female':0}
    equipa = {}
    equipg = {}
    equip={}
    equipv = {}
    equipc = {}
    equips = {}
    age=[]
    path='f'
    id=[]
    for line in path_list:
        try:
            file_path = line.strip('\n')  # 直接将文件中按行读到list里，效果与方法2一样
            if path==line.split('IM')[0]:
                # list.append(line)
                continue
            else:
                path = line.split('IM')[0]
                dsA = pydicom.dcmread(file_path, force=True)  # 读取头文件
                ag = int(dsA.PatientAge[0:3])  # PatientAge
                age.append(ag)
                # dsA.BodyPartExamined

                aa = dsA.Manufacturer.split(' ')[0]
                av = dsA.KVP
                ac=dsA.XRayTubeCurrent
                sl=dsA.SliceThickness
                if dsA.PatientID not in id:
                    id.append(dsA.PatientID)

                if aa + 'M' not in equipg and dsA.PatientSex == 'M':
                    equipg.setdefault(aa.split(' ')[0] + 'M', 1)
                elif aa + 'F' not in equipg and dsA.PatientSex == 'F':
                    equipg.setdefault(aa.split(' ')[0] + 'F', 1)
                elif dsA.PatientSex == 'F':
                        equipg[aa+'F'] += 1
                else:
                        equipg[aa+'M'] += 1

                if aa not in equip:
                    # 'GE MEDICAL SYSTEMS'
                    equipa.setdefault(aa.split(' ')[0], []).append(int(ag))
                    equip.setdefault(aa.split(' ')[0], 1)  # 'GE MEDICAL SYSTEMS'
                    equipv.setdefault(aa.split(' ')[0], []).append(int(av))
                    equipc.setdefault(aa.split(' ')[0],[]).append(int(ac))
                    equips.setdefault(aa.split(' ')[0], []).append(int(sl))
                else:
                    equipa[aa].append(int(ag))
                    equip[aa] += 1
                    equipc[aa].append(int(ac))
                    equipv[aa].append(int(av))
                    equips[aa].append(int(sl))
        except:
            continue


    agee=np.array(age)
    max=np.max(agee)
    min=np.min(agee)
    f1.writelines('slices numbers:'+str(len(path_list))+"\n")
    # f1.writelines('average age:' + str(min+(max-min)/2) +'  difference age:'+str((max-min)/2) +"\n")
    res = np.percentile(agee, (25, 50, 75), method='midpoint')
    f1.writelines('average age:' + str(res)+"\n")
    f1.writelines('male :'+str(gender['male']) +'female :'+str(gender['female'])+"\n")
    f1.writelines('patients :' + str(len(id)) + "\n")
    for eq in equipa:
        resa = np.percentile(equipa[eq], (25, 50, 75), method='midpoint')
        f1.writelines(eq+'age:'+str(resa) + "\n")

    for eq in equipg:
        resg = np.sum(equipg[eq])
        f1.writelines(eq+':'+str(resg) + "\n")

    for eq in equip:
        f1.writelines(eq+' equip:'+str(equip[eq]) + "\n")
    for eq in equipv:
        resv = np.percentile(equipv[eq], (25, 50, 75), method='midpoint')
        f1.writelines(eq+'voltage:'+str(resv) + "\n")
    for eq in equipc:
        resc = np.percentile(equipc[eq], (25, 50, 75), method='midpoint')
        f1.writelines(eq+' current:'+str(resc) + "\n")
    for eq in equips:
        ress = np.percentile(equips[eq], (25, 50, 75), method='midpoint')
        f1.writelines(eq+' slices:'+str(ress) + "\n")
    f1.close()#{'GE': 54900, 'SIEMENS': 18705, 'Philips': 1491}
    print("finished!")

# 将动脉血流，勾画的软斑块
def rmake_mask():
    # 保留动脉血流范围的图像；mask分别为动脉血流、钙化斑块、血栓、软斑块；如果只有动脉血流则保持不变。
    # path = "../../../data/p2_nii/jfj/p1/dis/dmzyyh/xb/DICOM3a/PA7" #
    path = "/media/bit301/data/yml/data/p2_nii/external/hnnk/dis/dml/PA4/SE0"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0
    for se2output in path_list:
        se0output = se2output.replace("2.nii.gz", "0.nii.gz")
        read = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read)
        img_array_w0 = to_windowdata(img_array0, 130, 10)
        img_array_w0 = np.where(img_array_w0 > 0, 1, 0)  # 支架与钙化斑块
        # img_array_w00=1-img_array_w0

        se1output = se2output.replace("2.nii.gz", "1.nii.gz")
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)

        read = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array2 = sitk.GetArrayFromImage(read)
        if img_array2.all() == 1:  # mask数值都为1，表示为只有血流。
            print("it's ok")
        else:  # 否者交换mask序号
            img_array2 = np.where(img_array2 == 1, 22, img_array2)
            img_array2 = np.where(img_array2 == 2, 1, img_array2)
            img_array2 = np.where(img_array2 == 22, 2, img_array2)
        img_array22 = np.where(img_array2 > 0, 1, 0)  # 将mask合并
        # 裁剪有效范围
        # 裁剪有效范围
        su = np.sum(img_array2, axis=(1, 2))
        index = np.where(su > 16)[0]  # 避免噪声影响
        # index=np.nonzero(img_array2)[0]
        st = min(index)
        en = max(index) + 1

        se3output = se2output.replace("2.nii.gz", "3.nii.gz")  # 血液分割标签。有可能不存在
        try:
            img3 = sitk.ReadImage(se3output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            img_array3 = sitk.GetArrayFromImage(img3)
            img_array33 = np.where(img_array3 > 0, 1, 0)
            # print(img_array2.min())
            # print(img_array2.max())

            img_array33 = img_array33 - img_array22  # 血流、化斑块与血栓、软斑块不重叠
            img_array33 = np.where(img_array33 > 0, 1, 0)
            img_array333 = img_array3 * img_array33 + 2  # 血栓和软斑块的mask
            img_array333 = np.where(img_array333 < 3, 0, img_array333)  # 血栓为3，斑块为4
            img_array333 = img_array2 + img_array333  # mask整合

            mask1 = img_array_w0 - img_array_w0 * img_array22  # 支架
            mask2 = 1 - mask1  # 去除金属支架
            img_array333 = mask2 * img_array333
            out2 = sitk.GetImageFromArray(img_array333[st:en, :, :].astype(np.int16))
            os.remove(se2output)
            os.remove(se3output)
            sitk.WriteImage(out2, se2output)
        except:
            print("不存在血栓和软斑块", se3output)
            out2 = sitk.GetImageFromArray(img_array2[st:en, :, :].astype(np.int16))
            os.remove(se2output)
            sitk.WriteImage(out2, se2output)
            # continue

        os.remove(se0output)
        os.remove(se1output)
        out0 = sitk.GetImageFromArray(img_array0[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out0, se0output)
        out1 = sitk.GetImageFromArray(img_array1[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out1, se1output)

        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


def crop_by_mask():
    # 保留动脉血流范围的图像；mask分别为动脉血流、钙化斑块、血栓、软斑块；如果只有动脉血流则保持不变。
    path = "../../../data/p2_s"  #
    # path = "/media/bit301/data/yml/data/p2_nii/external/lz/nor/PA192"
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0
    for se2output in path_list:
        se0output = se2output.replace("2.nii.gz", "0.nii.gz")
        read = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read)
        img_array_w0 = to_windowdata(img_array0, 130, 10)
        img_array_w0 = np.where(img_array_w0 > 0, 1, 0)  # 支架与钙化斑块
        # img_array_w00=1-img_array_w0

        se1output = se2output.replace("2.nii.gz", "1.nii.gz")
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)

        read = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array2 = sitk.GetArrayFromImage(read)
        # 裁剪有效范围
        su = np.sum(img_array2, axis=(1, 2))
        index = np.where(su > 16)[0]  # 避免一些小点出现
        # index=np.nonzero(img_array2)
        st = min(index)
        en = max(index) + 1

        # st=img_array1.shape[0]-img_array2.shape[0]#处理特殊情况
        # en=img_array1.shape[0]+1

        out2 = sitk.GetImageFromArray(img_array2[st:en, :, :].astype(np.int16))
        os.remove(se2output)
        sitk.WriteImage(out2, se2output)

        os.remove(se0output)
        os.remove(se1output)
        out0 = sitk.GetImageFromArray(img_array0[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out0, se0output)
        out1 = sitk.GetImageFromArray(img_array1[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out1, se1output)

        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)


def get_files_list():
    path = "/media/bit301/data/yml/data/p2_nii/external/"  # CT_CTA internal external
    files_list = "test.txt"
    f1 = open(files_list, "w")  # 564
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        # if "xm" in root or "p1/dmzyyh" in root or "hnnk" in root:
        #     if '0.nii.gz' in files:
        #         path_list.append(root.split("p2_nii")[1])
        for file in files:
            path = os.path.join(root, file)
            # if "xm" in path and "0.nii.gz" in path:
            if "0.nii.gz" in path:
                path_list.append(path)
    path_list = natsorted(path_list)
    for j in range(len(path_list)):
        f1.writelines(path_list[j] + "\n")
    # ff.close()
    f1.close()  # 关


# make the list of nnunet to 3DUxnet
def json_to_list():
    import json
    five_fold = "/home/wangtao/yml/project/python39/p3/nnUNet/DATASET/nnUNet_preprocessed/Dataset301_Aorta/splits_final.json"
    with open(five_fold, 'r', encoding='UTF-8') as f:
        files = json.load(f)

    root_and_json = "/home/wangtao/yml/project/python39/p3/nnUNet/DATASET/nnUNet_raw/Dataset301_Aorta/train.json"
    for i in range(len(files)):
        train_files = files[i]['train']
        val_files = files[i]['val']
        train = "train" + str(i) + ".txt"
        val = "val" + str(i) + ".txt"
        f1 = open(train, "w")  # 564
        f2 = open(val, "w")  # 188

        with open(root_and_json, encoding='utf-8') as f:
            while True:
                line = f.readline()
                if not line:  # 到 EOF，返回空字符串，则终止循环
                    break
                js = json.loads(line)
                for k, v in js.items():
                    if k in train_files:
                        f1.writelines(v + "\n")
                    else:
                        f2.writelines(v + "\n")
        f1.close()
        f2.close()

def radiologist_score_file():
    import openpyxl
    data = openpyxl.load_workbook('demo1.xlsx')
    sheetnames = data.get_sheet_names()
    table = data.active
    print(table.title)  # 输出表名
    path = "/media/bit301/data/yml/data/p2_nii/external/hnnk/"  # CT_CTA internal
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path:
                path_list.append(path)

    # f = open("hnnk.txt")  # test train！！  p_test1
    # path_list = []
    # for line in f.readlines():
    #     if "0.nii.gz" in line:
    #         path_list.append(line)

    path_list = natsorted(path_list)
    ii = 0
    nrows = 1
    for sub_path in path_list:
        write_path = sub_path.split('external')[1]
        table.cell(nrows + 1, 3).value = write_path
        nrows = nrows + 1
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    data.save('hnnk_MedNeXtx1c.xlsx')  #
    print("finished!")

def reshape_mask():
    path = "/media/bit301/data/yml/data/p2_nii/test/SwinUNETR/"#3DUXNET SwinUNETR
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                path_list.append(path)
    i = 0
    for file_path in path_list:
        img = sitk.ReadImage(file_path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array = sitk.GetArrayFromImage(img)
        if img_array.shape[1] != img_array.shape[2]:#target is zyx
            img_array = img_array.transpose((2, 1, 0))
            out = sitk.GetImageFromArray(img_array.astype(np.int16))
            sitk.WriteImage(out, file_path)
            i = i + 1
            if i % 10 == 0:
                print('numbers:', i)

def postpossess():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/test.txt")  # hnnk test.txt
    path_list = []
    ij = 0
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        if ij<51:
            ij=ij+1
            continue
            # break
        # path_list.append(line.split('\n')[0])
        path="/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA32/0.nii.gz"
        # path = line.split('\n')[0]
        img = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array = sitk.GetArrayFromImage(img)
        mask_path = path.replace("0.nii.gz", "2.nii.gz")
        mask_path1=mask_path
        # mask_path1=mask_path.replace("data/yml/data/p2_nii","use/p2")
        mask_read = sitk.ReadImage(mask_path1, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        mask_array = sitk.GetArrayFromImage(mask_read)
        mask=copy.deepcopy(mask_array)
        #process
        mask[mask > 0]=1
        img_array1=img_array*mask#勾画区域
        # img_array1[img_array1>130]=1#钙化
        img_array1=np.where(img_array1>130,1,0)
        inverted_mask=1-img_array1
        # img_array1 = (img_array1 > 130).astype(np.uint8)
        # inverted_mask = ~(img_array1.astype(bool)).astype(np.uint8)
        mask_array[mask_array==2]=1#让钙化也先划分为管腔
        mask_array=mask_array*inverted_mask#腾出钙化区域
        mask_array=mask_array+img_array1*2 #钙化为2
        if mask_array.max()>4:
            print(mask_path)
        label = sitk.GetImageFromArray(mask_array.astype(np.int16))
        sitk.WriteImage(label, mask_path)
        mask_array=0

        #model-mask
        mask_path = path.replace("external", "test/MedNeXtx2")  # lo:97  x1:54  xlo2:10
        mask_path = mask_path.replace("0.nii.gz", "2.nii.gz")
        mask_read = sitk.ReadImage(mask_path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        mask_array = sitk.GetArrayFromImage(mask_read)
        mask2 = copy.deepcopy(mask_array)
        # process
        mask2[mask2 > 0] = 1
        img_array2 = img_array * mask2
        img_array2=np.where(img_array2>130,1,0)
        inverted_mask=1-img_array2
        mask_array[mask_array == 2] = 1  # 让钙化也先划分为管腔
        mask_array = mask_array * inverted_mask  # 腾出钙化区域
        mask_array = mask_array + img_array2*2 #钙化为2
        if mask_array.max()>4:
            print(mask_path)
        label = sitk.GetImageFromArray(mask_array.astype(np.int16))
        out_path=mask_path.replace("MedNeXtx2","MedNeXtx22")
        out_put = out_path.split("2.nii.gz")[0]
        if not os.path.isdir(out_put):
            os.makedirs(out_put)
        # sitk.WriteImage(label, out_path)
        ij = ij + 1
        if ij % 10 == 0:
            print('numbers:', ij)
    print('finished:!')

def remove_case():
    f = open("/media/bit301/data/yml/project/python39/p2/Aorta_net/data/test.txt")  # hnnk test.txt
    path_list = []
    ij = 0
    min=500
    for line in f.readlines():  # tile_step_size=0.75较好处理官腔错位问题
        # if ij < 51:
        #     ij = ij + 1
        #     continue
        #     # break
        # path_list.append(line.split('\n')[0])
        # path="/media/bit301/data/yml/data/p2_nii/external/cq/dis/dmzyyh/PA177/0.nii.gz"
        path = line.split('\n')[0]
        img = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array = sitk.GetArrayFromImage(img)
        if img_array.shape[0]<min:
            min=img_array.shape[0]
            # print(path)
    print(min)

def postpossess2():
    path="/media/bit301/data/yml/data/p2_nii/external/lz/dis/dmzyyh/PA32/2.nii.gz"

if __name__ == '__main__':
    # copy_and_paste()
    # copy_xz_paste()
    # remove_file()
    # mv_file()
    # make_process()#处理301数据
    # make_process_xy()#处理xy数据
    # make_abd_process5()#cq
    # make_abd_process5b()#配准
    # make_abd_process6()
    # nii_to_image()
    # display()

    # make_mask()
    # get_abdtest_list()
    # check_slice_num()
    # split()
    # check_image_label()
    # vessel_to_Calcified()
    # vessel_to_Calcified_test()
    # remove_s_slices()
    # remove_blank_slices()
    # types_statistic()
    # get_slices()
    # statistic()
    # rmake_mask()
    # crop_by_mask()
    # json_to_list()
    # get_files_list()
    # radiologist_score_file()
    # reshape_mask()
    # cq_statistic()
    # lz_statistic()
    postpossess()
    # remove_case()
    a = 1
